"""
Performance Management System - Main Application
A comprehensive PMS built with Streamlit and Supabase
"""

import sys
sys.setrecursionlimit(3000)
import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import pytz
import plotly.express as px
import plotly.graph_objects as go
import calendar
import io
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import numpy as np
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from dateutil.relativedelta import relativedelta
from database import get_supabase_client
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import secrets
import os
from datetime import datetime, timedelta
import streamlit.components.v1 as components

st.markdown("""
<style>

.q-btn-wrapper div[data-testid="baseButton-secondary"] {
    background-color: #475569; /* Slate grey */
    color: white !important;
    border-radius: 8px !important;
    border: 1px solid #334155 !important;
    padding: 10px 14px !important;
    font-weight: 600 !important;
    transition: 0.2s ease-in-out;
}

.q-btn-wrapper div[data-testid="baseButton-secondary"]:hover {
    background-color: #334155 !important;
}

</style>
""", unsafe_allow_html=True)



def check_password_strength(password):
    """Check password strength and return score, color, and feedback"""
    if not password:
        return 0, "#94a3b8", "No password entered", []
    
    score = 0
    feedback = []
    
    # Length check
    if len(password) >= 8:
        score += 25
    elif len(password) >= 6:
        score += 15
        feedback.append("🔸 Use at least 8 characters")
    else:
        feedback.append("❌ Too short (minimum 6 characters)")
    
    # Uppercase check
    if any(c.isupper() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add uppercase letters (A-Z)")
    
    # Lowercase check
    if any(c.islower() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add lowercase letters (a-z)")
    
    # Number check
    if any(c.isdigit() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add numbers (0-9)")
    
    # Special character check
    special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if any(c in special_chars for c in password):
        score += 15
    else:
        feedback.append("🔸 Add special characters (!@#$%)")
    
    # Determine strength level and color
    if score >= 85:
        strength = "Very Strong 💪"
        color = "#10b981"  # Green
    elif score >= 70:
        strength = "Strong 👍"
        color = "#22c55e"  # Light green
    elif score >= 50:
        strength = "Medium ⚡"
        color = "#f59e0b"  # Orange
    elif score >= 30:
        strength = "Weak ⚠️"
        color = "#fb923c"  # Light orange
    else:
        strength = "Very Weak ❌"
        color = "#ef4444"  # Red
    
    return score, color, strength, feedback

def calculate_performance_metrics(goals):
    """Calculate comprehensive performance metrics"""
    if not goals:
        return None
    
    total_goals = len(goals)
    completed = len([g for g in goals if g.get('status') == 'Completed'])
    active = len([g for g in goals if g.get('status') == 'Active'])
    on_hold = len([g for g in goals if g.get('status') == 'On Hold'])
    cancelled = len([g for g in goals if g.get('status') == 'Cancelled'])
    
    # Calculate average progress
    total_progress = 0
    goal_count = 0
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        total_progress += progress
        goal_count += 1
    
    avg_progress = total_progress / goal_count if goal_count > 0 else 0
    
    # Calculate completion rate
    completion_rate = (completed / total_goals * 100) if total_goals > 0 else 0
    
    # Calculate on-time completion rate
    on_time = 0
    overdue = 0
    today = date.today()
    
    for goal in goals:
        if goal.get('status') == 'Completed':
            end_date = datetime.strptime(str(goal.get('end_date')), '%Y-%m-%d').date()
            if end_date >= today:
                on_time += 1
            else:
                overdue += 1
    
    on_time_rate = (on_time / completed * 100) if completed > 0 else 0
    
    return {
        'total_goals': total_goals,
        'completed': completed,
        'active': active,
        'on_hold': on_hold,
        'cancelled': cancelled,
        'avg_progress': avg_progress,
        'completion_rate': completion_rate,
        'on_time': on_time,
        'overdue': overdue,
        'on_time_rate': on_time_rate
    }


def get_trend_data(user_id, months=6):
    """Get performance trend data for last N months"""
    trends = []
    today = date.today()
    
    for i in range(months):
        month_date = today - relativedelta(months=i)
        year = month_date.year
        month = month_date.month
        quarter = ((month - 1) // 3) + 1
        
        # Get goals for this month
        try:
            goals = db.get_month_goals(user_id, year, quarter, month)
            
            if goals and len(goals) > 0:
                metrics = calculate_performance_metrics(goals)
                trends.append({
                    'month': month_date.strftime('%b %Y'),
                    'completion_rate': metrics['completion_rate'],
                    'avg_progress': metrics['avg_progress'],
                    'total_goals': metrics['total_goals'],
                    'completed': metrics['completed']
                })
        except Exception as e:
            # Skip months with errors
            print(f"Error getting data for {month_date}: {str(e)}")
            continue
    
    return list(reversed(trends))


# ============================================
# ADVANCED CHARTS
# ============================================

def create_performance_gauge(value, title="Performance"):
    """Create a gauge chart for performance"""
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=value,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title, 'font': {'size': 24}},
        delta={'reference': 75, 'increasing': {'color': "green"}},
        gauge={
            'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
            'bar': {'color': "darkblue"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, 30], 'color': '#fee2e2'},
                {'range': [30, 60], 'color': '#fed7aa'},
                {'range': [60, 80], 'color': '#fef3c7'},
                {'range': [80, 100], 'color': '#d1fae5'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    
    fig.update_layout(
        height=300,
        margin=dict(l=20, r=20, t=60, b=20),
        paper_bgcolor="white",
        font={'color': "darkblue", 'family': "Arial"}
    )
    
    return fig


def create_trend_chart(trend_data):
    """Create performance trend line chart"""
    if not trend_data:
        return None
    
    df = pd.DataFrame(trend_data)
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Goal Completion Trend', 'Average Progress Trend'),
        vertical_spacing=0.15
    )
    
    # Completion rate trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['completion_rate'],
            mode='lines+markers',
            name='Completion Rate',
            line=dict(color='#3b82f6', width=3),
            marker=dict(size=8)
        ),
        row=1, col=1
    )
    
    # Average progress trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['avg_progress'],
            mode='lines+markers',
            name='Avg Progress',
            line=dict(color='#10b981', width=3),
            marker=dict(size=8)
        ),
        row=2, col=1
    )
    
    fig.update_xaxes(title_text="Month", row=2, col=1)
    fig.update_yaxes(title_text="Completion %", row=1, col=1)
    fig.update_yaxes(title_text="Progress %", row=2, col=1)
    
    fig.update_layout(
        height=600,
        showlegend=True,
        title_text="Performance Trends (Last 6 Months)",
        title_font_size=20
    )
    
    return fig


def create_status_distribution_chart(goals):
    """Create pie chart for goal status distribution"""
    status_counts = {}
    for goal in goals:
        status = goal.get('status', 'Active')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    colors_map = {
        'Active': '#3b82f6',
        'Completed': '#10b981',
        'On Hold': '#f59e0b',
        'Cancelled': '#ef4444'
    }
    
    fig = go.Figure(data=[go.Pie(
        labels=list(status_counts.keys()),
        values=list(status_counts.values()),
        hole=0.4,
        marker_colors=[colors_map.get(k, '#64748b') for k in status_counts.keys()]
    )])
    
    fig.update_layout(
        title_text="Goal Status Distribution",
        title_font_size=20,
        height=400,
        showlegend=True
    )
    
    return fig


def create_vertical_performance_chart(goals):
    """Create bar chart for performance by vertical"""
    vertical_data = {}
    
    for goal in goals:
        vertical = goal.get('vertical', 'Unassigned')
        if vertical not in vertical_data:
            vertical_data[vertical] = {'total': 0, 'completed': 0, 'progress': []}
        
        vertical_data[vertical]['total'] += 1
        if goal.get('status') == 'Completed':
            vertical_data[vertical]['completed'] += 1
        
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        vertical_data[vertical]['progress'].append(progress)
    
    verticals = list(vertical_data.keys())
    completion_rates = [
        (vertical_data[v]['completed'] / vertical_data[v]['total'] * 100) 
        if vertical_data[v]['total'] > 0 else 0 
        for v in verticals
    ]
    avg_progress = [
        sum(vertical_data[v]['progress']) / len(vertical_data[v]['progress']) 
        if vertical_data[v]['progress'] else 0 
        for v in verticals
    ]
    
    fig = go.Figure(data=[
        go.Bar(name='Completion Rate', x=verticals, y=completion_rates, marker_color='#3b82f6'),
        go.Bar(name='Avg Progress', x=verticals, y=avg_progress, marker_color='#10b981')
    ])
    
    fig.update_layout(
        barmode='group',
        title_text="Performance by Vertical",
        title_font_size=20,
        xaxis_title="Vertical",
        yaxis_title="Percentage (%)",
        height=400
    )
    
    return fig


def create_heatmap_calendar(goals, year, month):
    """Create calendar heatmap for daily goal achievements"""
    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Initialize achievement data
    daily_achievements = {}
    for day in range(1, days_in_month + 1):
        daily_achievements[day] = 0
    
    # Calculate daily achievements
    for goal in goals:
        for week in range(1, 5):
            week_achievement = goal.get(f'week{week}_achievement', 0)
            if week_achievement > 0:
                # Distribute across week days
                week_start, week_end = get_week_dates(year, month, week)
                days_in_week = (week_end - week_start).days + 1
                daily_avg = week_achievement / days_in_week
                
                for day in range(week_start.day, min(week_end.day + 1, days_in_month + 1)):
                    daily_achievements[day] += daily_avg
    
    # Create matrix for heatmap
    weeks = 5
    days = 7
    matrix = np.zeros((weeks, days))
    day_labels = []
    
    first_day = date(year, month, 1)
    first_weekday = first_day.weekday()
    
    day_counter = 1
    for week in range(weeks):
        for day in range(days):
            if week == 0 and day < first_weekday:
                matrix[week][day] = np.nan
            elif day_counter > days_in_month:
                matrix[week][day] = np.nan
            else:
                matrix[week][day] = daily_achievements.get(day_counter, 0)
                day_counter += 1
    
    fig = go.Figure(data=go.Heatmap(
        z=matrix,
        x=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        y=[f'Week {i+1}' for i in range(weeks)],
        colorscale='Blues',
        showscale=True
    ))
    
    fig.update_layout(
        title=f"Daily Achievement Heatmap - {get_month_name(month)} {year}",
        title_font_size=20,
        height=400
    )
    
    return fig


# ============================================
# PDF REPORT GENERATION
# ============================================

def generate_performance_report_pdf(user, goals, period="Monthly"):
    """Generate comprehensive PDF performance report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph(f"Performance Report - {period}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # User Information
    story.append(Paragraph("Employee Information", heading_style))
    user_data = [
        ['Name:', user['name']],
        ['Email:', user['email']],
        ['Role:', user['role']],
        ['Department:', user.get('department', 'N/A')],
        ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    user_table = Table(user_data, colWidths=[2*inch, 4*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(user_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Performance Metrics
    metrics = calculate_performance_metrics(goals)
    
    if metrics:
        story.append(Paragraph("Performance Summary", heading_style))
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Goals', str(metrics['total_goals'])],
            ['Completed Goals', str(metrics['completed'])],
            ['Active Goals', str(metrics['active'])],
            ['Completion Rate', f"{metrics['completion_rate']:.1f}%"],
            ['Average Progress', f"{metrics['avg_progress']:.1f}%"],
            ['On-Time Completion Rate', f"{metrics['on_time_rate']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Goals Details
    story.append(Paragraph("Goals Details", heading_style))
    
    goals_data = [['Goal Title', 'Vertical', 'Target', 'Achievement', 'Progress', 'Status']]
    
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        
        goals_data.append([
            goal['goal_title'][:30],
            goal.get('vertical', 'N/A'),
            str(goal.get('monthly_target', 0)),
            str(goal.get('monthly_achievement', 0)),
            f"{progress:.1f}%",
            goal.get('status', 'Active')
        ])
    
    goals_table = Table(goals_data, colWidths=[2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    goals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(goals_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


# ============================================
# ANALYTICS PAGE
# ============================================

def display_analytics_page():
    """Display comprehensive analytics page"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()
    role = user['role']
    
    st.title("📊 Advanced Analytics & Reports")
    
    # Filters
    col_filter1, col_filter2, col_filter3 = st.columns(3)
    
    with col_filter1:
        if role == 'HR':
            all_users = db.get_all_users()
            selected_user = st.selectbox(
                "Select User",
                ["My Analytics"] + [f"{u['name']} ({u['email']})" for u in all_users]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(u for u in all_users if u['email'] == user_email)
        elif role == 'Manager':
            team_members = db.get_team_members(user['id'])
            selected_user = st.selectbox(
                "Select Team Member",
                ["My Analytics"] + [f"{m['name']} ({m['email']})" for m in team_members]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(m for m in team_members if m['email'] == user_email)
        else:
            analysis_user = user
    
    with col_filter2:
        analysis_period = st.selectbox(
            "Analysis Period",
            ["Current Month", "Current Quarter", "Current Year", "Last 6 Months", "All Time"]
        )
    
    with col_filter3:
        view_type = st.selectbox(
            "View Type",
            ["Overview", "Trends", "Comparisons", "Detailed"]
        )
    
    # Get goals based on period
    all_goals = db.get_user_all_goals(analysis_user['id'])
    
    if analysis_period == "Current Month":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year and g.get('month') == today.month]
    elif analysis_period == "Current Quarter":
        today = date.today()
        quarter = ((today.month - 1) // 3) + 1
        goals = [g for g in all_goals if g['year'] == today.year and g.get('quarter') == quarter]
    elif analysis_period == "Current Year":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year]
    else:
        goals = all_goals
    
    if not goals:
        st.info(f"No goals found for {analysis_user['name']} in selected period")
        return
    
    st.markdown("---")
    
    # Calculate metrics
    metrics = calculate_performance_metrics(goals)
    
    # Overview Tab
    if view_type == "Overview":
        st.subheader(f"Performance Overview - {analysis_user['name']}")
        
        # KPI Cards
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            render_metric_card("Total Goals", str(metrics['total_goals']), color="#3b82f6")
        with col2:
            render_metric_card("Completed", str(metrics['completed']), color="#10b981")
        with col3:
            render_metric_card("Active", str(metrics['active']), color="#f59e0b")
        with col4:
            render_metric_card("Completion Rate", f"{metrics['completion_rate']:.1f}%", color="#8b5cf6")
        with col5:
            render_metric_card("Avg Progress", f"{metrics['avg_progress']:.1f}%", color="#ec4899")
        
        st.markdown("---")
        
        # Charts Row 1
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            # Performance Gauge
            gauge_fig = create_performance_gauge(metrics['avg_progress'], "Overall Performance")
            st.plotly_chart(gauge_fig, use_container_width=True)
        
        with col_chart2:
            # Status Distribution
            status_fig = create_status_distribution_chart(goals)
            st.plotly_chart(status_fig, use_container_width=True)
        
        st.markdown("---")
        
        # Charts Row 2
        col_chart3, col_chart4 = st.columns(2)
        
        with col_chart3:
            # Vertical Performance
            vertical_fig = create_vertical_performance_chart(goals)
            st.plotly_chart(vertical_fig, use_container_width=True)
        
        with col_chart4:
            # On-time vs Overdue
            on_time_data = pd.DataFrame({
                'Category': ['On-Time', 'Overdue'],
                'Count': [metrics['on_time'], metrics['overdue']]
            })
            fig_ontime = px.bar(on_time_data, x='Category', y='Count',
                               title="On-Time vs Overdue Completion",
                               color='Category',
                               color_discrete_map={'On-Time': '#10b981', 'Overdue': '#ef4444'})
            fig_ontime.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig_ontime, use_container_width=True)
    
    # Trends Tab
    elif view_type == "Trends":
        st.subheader(f"Performance Trends - {analysis_user['name']}")
        
        trend_data = get_trend_data(analysis_user['id'], months=6)
        
        if trend_data and len(trend_data) > 0:
            # Main trend chart
            trend_fig = create_trend_chart(trend_data)
            st.plotly_chart(trend_fig, use_container_width=True)
            
            st.markdown("---")
            
            # Monthly Achievement Trend
            st.subheader("Monthly Goal Achievement Trend")
            
            monthly_data = []
            for month_info in trend_data:
                monthly_data.append({
                    'Month': month_info['month'],
                    'Goals': month_info['total_goals'],
                    'Completed': month_info['completed']
                })
            
            df_monthly = pd.DataFrame(monthly_data)
            
            if not df_monthly.empty:
                fig_monthly = px.line(df_monthly, x='Month', y=['Goals', 'Completed'],
                                     title="Goals Created vs Completed",
                                     markers=True)
                fig_monthly.update_layout(height=400)
                st.plotly_chart(fig_monthly, use_container_width=True)
            else:
                st.info("No monthly data available for visualization")
        else:
            st.info("Not enough data for trend analysis. Please ensure you have goals spanning multiple months.")
    
    # Comparisons Tab
    elif view_type == "Comparisons":
        st.subheader(f"🔄 Performance Comparisons - {analysis_user['name']}")
        
        if role in ['HR', 'Manager']:
            # Get comparison users based on the selected user's role
            if role == 'HR':
                # HR: Compare with same role users
                selected_role = analysis_user['role']
                all_users = db.get_all_users()
                
                if selected_role == 'Manager':
                    compare_users = [u for u in all_users if u['role'] == 'Manager']
                    comparison_title = "Compare with Other Managers"
                elif selected_role == 'Employee':
                    compare_users = [u for u in all_users if u['role'] == 'Employee']
                    comparison_title = "Compare with Other Employees"
                else:  # HR viewing another HR
                    compare_users = [u for u in all_users if u['role'] == 'HR']
                    comparison_title = "Compare with Other HR Members"
                    
            else:  # Manager
                # Manager compares team members only (not themselves)
                if analysis_user['id'] == user['id']:
                    # Manager viewing their own analytics - compare with team only
                    compare_users = db.get_team_members(user['id'])
                    comparison_title = "Compare Team Members"
                else:
                    # Manager viewing a team member - compare with other team members
                    team_members = db.get_team_members(user['id'])
                    compare_users = [m for m in team_members]
                    comparison_title = "Compare with Team Members"
            
            if compare_users:
                # Show appropriate info message
                if role == 'Manager' and analysis_user['id'] == user['id']:
                    st.info(f"**{comparison_title}** ({len(compare_users)} team members)\n\n💡 *Note: You are viewing your team's performance comparison (manager is not included in rankings)*")
                else:
                    st.info(f"**{comparison_title}** ({len(compare_users)} users)")
                
                comparison_data = []
                
                # Add all users (including current)
                for comp_user in compare_users:
                    comp_goals = db.get_user_all_goals(comp_user['id'])
                    if comp_goals:
                        comp_metrics = calculate_performance_metrics(comp_goals)
                        is_current = (comp_user['id'] == analysis_user['id'])
                        comparison_data.append({
                            'Name': f" {comp_user['name']}" if is_current else comp_user['name'],
                            'Display_Name': comp_user['name'],
                            'Total Goals': comp_metrics['total_goals'],
                            'Completion Rate': comp_metrics['completion_rate'],
                            'Avg Progress': comp_metrics['avg_progress'],
                            'Is_Current': is_current,
                            'User_ID': comp_user['id']
                        })
                
                if comparison_data and len(comparison_data) > 0:
                    df_compare = pd.DataFrame(comparison_data)
                    
                    # Highlight current user
                    st.success(f" **{analysis_user['name']}** is highlighted with ⭐ in the comparisons below")
                    
                    # Comparison bar chart
                    fig_compare = px.bar(
                        df_compare, 
                        x='Name', 
                        y=['Completion Rate', 'Avg Progress'],
                        title=f"{comparison_title} - Performance Metrics",
                        barmode='group',
                        height=500,
                        color_discrete_map={
                            'Completion Rate': '#3b82f6',
                            'Avg Progress': '#10b981'
                        }
                    )
                    fig_compare.update_layout(
                        xaxis_tickangle=-45,
                        xaxis_title="",
                        yaxis_title="Percentage (%)"
                    )
                    st.plotly_chart(fig_compare, use_container_width=True)
                    
                    st.markdown("---")
                    
                    # Comparison table with ranking
                    st.subheader(" Detailed Comparison Table")
                    
                    # Sort by completion rate for ranking
                    df_sorted = df_compare.sort_values('Completion Rate', ascending=False).reset_index(drop=True)
                    df_sorted.insert(0, 'Rank', range(1, len(df_sorted) + 1))
                    
                    # Create display dataframe
                    df_display = df_sorted[['Rank', 'Name', 'Total Goals', 'Completion Rate', 'Avg Progress']].copy()
                    df_display['Completion Rate'] = df_display['Completion Rate'].apply(lambda x: f"{x:.1f}%")
                    df_display['Avg Progress'] = df_display['Avg Progress'].apply(lambda x: f"{x:.1f}%")
                    
                    st.dataframe(df_display, use_container_width=True, height=400)
                    
                    # Performance insights
                    st.markdown("---")
                    st.subheader(" Performance Insights")
                    
                    # Calculate metrics correctly
                    avg_completion = df_compare['Completion Rate'].mean()
                    avg_progress = df_compare['Avg Progress'].mean()
                    
                    # Get current user's metrics
                    current_user_data = df_compare[df_compare['Is_Current'] == True]
                    
                    if not current_user_data.empty:
                        user_completion = current_user_data['Completion Rate'].iloc[0]
                        user_progress = current_user_data['Avg Progress'].iloc[0]
                        
                        # Get current user's rank
                        current_rank_row = df_sorted[df_sorted['Is_Current'] == True]
                        if not current_rank_row.empty:
                            current_rank = current_rank_row.iloc[0]['Rank']
                            
                            # Display rank
                            col_rank, col_space = st.columns([1, 3])
                            with col_rank:
                                st.metric("Current User's Rank", f"#{current_rank} out of {len(df_sorted)}")
                        
                        col_insight1, col_insight2 = st.columns(2)
                        
                        with col_insight1:
                            st.markdown("**Completion Rate Analysis:**")
                            diff = user_completion - avg_completion
                            if diff > 0:
                                st.success(f"**Above Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **+{diff:.1f}%** higher")
                            elif diff < 0:
                                st.warning(f"⚠️ **Below Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **{abs(diff):.1f}%** lower")
                            else:
                                st.info(f" **Average Performance**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nMatches group average: **{avg_completion:.1f}%**")
                        
                        with col_insight2:
                            st.markdown("**Progress Analysis:**")
                            prog_diff = user_progress - avg_progress
                            if prog_diff > 0:
                                st.success(f" **Above Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **+{prog_diff:.1f}%** higher")
                            elif prog_diff < 0:
                                st.warning(f"**Below Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **{abs(prog_diff):.1f}%** lower")
                            else:
                                st.info(f"**Average Performance**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nMatches group average: **{avg_progress:.1f}%**")
                        
                        # Top performer
                        st.markdown("---")
                        top_performer = df_sorted.iloc[0]
                        if top_performer['Is_Current']:
                            st.success(f" **Congratulations!** {analysis_user['name']} has the highest completion rate ({top_performer['Completion Rate']:.1f}%) in the group!")
                        else:
                            st.info(f"**Top Performer:** {top_performer['Display_Name']} leads with **{top_performer['Completion Rate']:.1f}%** completion rate")
                    
                else:
                    st.info("No comparison data available - users need to have goals to be included in comparison")
            else:
                st.info(f"No {comparison_title.lower()} available for comparison")
        else:
            st.warning(" Comparison view is only available for HR and Managers")
    
    # Detailed Tab
    elif view_type == "Detailed":
        st.subheader(f" Detailed Goal Analysis - {analysis_user['name']}")
        
        # Goals breakdown by status
        st.markdown("### Goals by Status")
        
        status_groups = {}
        for goal in goals:
            status = goal.get('status', 'Active')
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(goal)
        
        for status, status_goals in status_groups.items():
            with st.expander(f"{status} Goals ({len(status_goals)})"):
                for goal in status_goals:
                    progress = calculate_progress(
                        goal.get('monthly_achievement', 0),
                        goal.get('monthly_target', 1)
                    )
                    
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Vertical: {goal.get('vertical', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                    with col2:
                        st.metric("Target", goal.get('monthly_target', 0))
                    with col3:
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown("---")
        
        st.markdown("---")
        
        # Weekly breakdown
        st.markdown("### Weekly Performance Breakdown")
        
        weekly_data = {
            'Week 1': {'target': 0, 'achievement': 0},
            'Week 2': {'target': 0, 'achievement': 0},
            'Week 3': {'target': 0, 'achievement': 0},
            'Week 4': {'target': 0, 'achievement': 0}
        }
        
        for goal in goals:
            for week in range(1, 5):
                weekly_data[f'Week {week}']['target'] += goal.get(f'week{week}_target', 0)
                weekly_data[f'Week {week}']['achievement'] += goal.get(f'week{week}_achievement', 0)
        
        weekly_df = pd.DataFrame({
            'Week': list(weekly_data.keys()),
            'Target': [weekly_data[w]['target'] for w in weekly_data.keys()],
            'Achievement': [weekly_data[w]['achievement'] for w in weekly_data.keys()]
        })
        
        fig_weekly = px.bar(weekly_df, x='Week', y=['Target', 'Achievement'],
                           title="Weekly Targets vs Achievements",
                           barmode='group',
                           height=400)
        st.plotly_chart(fig_weekly, use_container_width=True)
    
    # Export Report
    st.markdown("---")
    st.subheader("Export Report")
    
    col_export1, col_export2, col_export3 = st.columns(3)
    
    with col_export1:
        if st.button("Export to PDF", use_container_width=True):
            with st.spinner("Generating PDF report..."):
                pdf_buffer = generate_performance_report_pdf(analysis_user, goals, analysis_period)
                st.download_button(
                    label=" Download PDF Report",
                    data=pdf_buffer,
                    file_name=f"Performance_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
    
    with col_export2:
        if st.button(" Export to Excel", use_container_width=True):
            # Create Excel with metrics
            excel_buffer = export_goals_to_excel(analysis_user['id'], 
                                                 goals[0]['year'] if goals else date.today().year,
                                                 goals[0].get('quarter', 1) if goals else 1,
                                                 goals[0].get('month', 1) if goals else 1)
            if excel_buffer:
                st.download_button(
                    label="Download Excel Report",
                    data=excel_buffer,
                    file_name=f"Goals_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
def get_completable_goals(user_id):
    """Get all goals that can be auto-completed"""
    all_goals = db.get_user_all_goals(user_id)
    
    completable = []
    for goal in all_goals:
        if goal.get('status') == 'Active':
            progress = calculate_progress(
                goal.get('monthly_achievement', 0),
                goal.get('monthly_target', 1)
            )
            if progress >= 100:
                completable.append(goal)
    
    return completable


def auto_complete_goal(goal_id, completed_by=None, remarks=None):
    """Auto-complete a goal that has reached 100%"""
    updates = {
        'status': 'Completed',
        'completed_at': datetime.now(IST).isoformat()
    }
    
    if completed_by:
        updates['completed_by'] = completed_by
    
    if remarks:
        updates['completion_remarks'] = remarks
    
    return db.update_goal(goal_id, updates)
def display_auto_complete_banner():
    """Show banner with auto-complete suggestions"""
    user = st.session_state.user
    
    completable_goals = get_completable_goals(user['id'])
    
    if completable_goals:
        with st.container():
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                <h3 style="color: white; margin: 0;">🎉 Goals Ready for Completion!</h3>
            </div>
            """, unsafe_allow_html=True)
            
            st.info(f"You have **{len(completable_goals)} goal(s)** at 100% progress")
            
            with st.expander(f"View {len(completable_goals)} Completable Goals", expanded=True):
                for goal in completable_goals:
                    col1, col2, col3 = st.columns([3, 1, 1])
                    
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Target: {goal.get('monthly_target', 0)}")
                    
                    with col2:
                        st.success("100% ")
                    
                    with col3:
                        if st.button("Complete", key=f"complete_{goal['goal_id']}", use_container_width=True):
                            if auto_complete_goal(goal['goal_id'], completed_by=user['id']):
                                st.success(f"Completed!")
                                st.rerun()
                
                st.markdown("---")
                
                col_bulk1, col_bulk2 = st.columns(2)
                
                with col_bulk1:
                    if st.button(" Complete All", type="primary", use_container_width=True):
                        for goal in completable_goals:
                            auto_complete_goal(goal['goal_id'], completed_by=user['id'])
                        st.success(f"Completed {len(completable_goals)} goals!")
                        st.balloons()
                        st.rerun()
                
                with col_bulk2:
                    if st.button("⏭Remind Later", use_container_width=True):
                        st.rerun()

def calculate_performance_score(stats):
    """Calculate weighted performance score for ranking"""
    # Weighted scoring system
    completion_weight = 0.4  # 40%
    progress_weight = 0.3    # 30%
    total_goals_weight = 0.2  # 20%
    on_time_weight = 0.1     # 10%
    
    completion_score = stats.get('completion_rate', 0) * completion_weight
    progress_score = stats.get('avg_progress', 0) * progress_weight
    goals_score = min(stats.get('total_goals', 0) * 5, 100) * total_goals_weight  # Cap at 100
    on_time_score = stats.get('on_time_rate', 0) * on_time_weight
    
    total_score = completion_score + progress_score + goals_score + on_time_score
    
    return round(total_score, 2)


def get_current_team_rankings(manager_id, year=None, month=None):
    """Get current team rankings with scores"""
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    team_members = db.get_team_members(manager_id)
    
    rankings = []
    for member in team_members:
        # Get goals for specific month
        quarter = ((month - 1) // 3) + 1
        goals = db.get_month_goals(member['id'], year, quarter, month)
        
        if goals:
            stats = calculate_performance_metrics(goals)
            score = calculate_performance_score(stats)
            
            rankings.append({
                'employee_id': member['id'],
                'name': member['name'],
                'email': member['email'],
                'designation': member.get('designation', 'N/A'),
                'department': member.get('department', 'N/A'),
                'total_goals': stats['total_goals'],
                'completed_goals': stats['completed'],
                'completion_rate': stats['completion_rate'],
                'avg_progress': stats['avg_progress'],
                'on_time_rate': stats.get('on_time_rate', 0),
                'score': score
            })
    
    # Sort by score (highest first)
    rankings.sort(key=lambda x: x['score'], reverse=True)
    
    # Assign ranks
    for idx, ranking in enumerate(rankings):
        ranking['rank'] = idx + 1
    
    return rankings


def save_monthly_rankings(manager_id, year, month):
    """Save monthly rankings to database"""
    rankings = get_current_team_rankings(manager_id, year, month)
    
    if not rankings:
        return False
    
    saved_count = 0
    for ranking in rankings:
        try:
            # Check if ranking already exists
            existing = supabase.table('team_rankings').select('id').eq(
                'manager_id', manager_id
            ).eq('employee_id', ranking['employee_id']).eq(
                'year', year
            ).eq('month', month).execute()
            
            ranking_data = {
                'manager_id': manager_id,
                'employee_id': ranking['employee_id'],
                'year': year,
                'month': month,
                'rank': ranking['rank'],
                'total_goals': ranking['total_goals'],
                'completed_goals': ranking['completed_goals'],
                'completion_rate': ranking['completion_rate'],
                'avg_progress': ranking['avg_progress'],
                'score': ranking['score']
            }
            
            if existing.data:
                # Update existing
                supabase.table('team_rankings').update(ranking_data).eq(
                    'id', existing.data[0]['id']
                ).execute()
            else:
                # Insert new
                supabase.table('team_rankings').insert(ranking_data).execute()
            
            saved_count += 1
        except Exception as e:
            print(f"Error saving ranking: {str(e)}")
            continue
    
    return saved_count > 0


def get_historical_rankings(manager_id, employee_id, months=6):
    """Get historical rankings for an employee"""
    try:
        result = supabase.table('team_rankings').select('*').eq(
            'manager_id', manager_id
        ).eq('employee_id', employee_id).order(
            'year', desc=True
        ).order('month', desc=True).limit(months).execute()
        
        return result.data if result.data else []
    except Exception as e:
        print(f"Error getting historical rankings: {str(e)}")
        return []


def get_average_ranking(manager_id, employee_id, months=6):
    """Calculate average ranking for an employee over last N months"""
    history = get_historical_rankings(manager_id, employee_id, months)
    
    if not history:
        return None
    
    total_rank = sum(r['rank'] for r in history)
    avg_rank = total_rank / len(history)
    
    return {
        'avg_rank': round(avg_rank, 1),
        'best_rank': min(r['rank'] for r in history),
        'worst_rank': max(r['rank'] for r in history),
        'months_tracked': len(history)
    }


# ============================================
# DISPLAY TEAM RANKINGS (REPLACE TEAM PERFORMANCE)
# ============================================

def display_team_rankings_dashboard(manager_id):
    """Display team rankings in manager dashboard"""
    st.subheader(" Team Performance Rankings")
    
    # Month selector
    col_month1, col_month2, col_month3 = st.columns([2, 1, 1])
    
    with col_month1:
        current_date = date.today()
        selected_year = st.selectbox(
            "Year",
            range(current_date.year, current_date.year - 3, -1),
            key="rank_year"
        )
    
    with col_month2:
        selected_month = st.selectbox(
            "Month",
            range(1, 13),
            index=current_date.month - 1,
            format_func=lambda x: get_month_name(x),
            key="rank_month"
        )
    
    with col_month3:
        if st.button(" Save Rankings", use_container_width=True, 
                    help="Save current month's rankings to history"):
            if save_monthly_rankings(manager_id, selected_year, selected_month):
                st.success(f" Rankings saved for {get_month_name(selected_month)} {selected_year}")
                st.rerun()
            else:
                st.error(" Failed to save rankings")
    
    # Get rankings
    rankings = get_current_team_rankings(manager_id, selected_year, selected_month)
    
    if not rankings:
        st.info(f"No performance data available for {get_month_name(selected_month)} {selected_year}")
        return
    
    st.markdown("---")
    
    # Top 3 Podium Display
    if len(rankings) >= 3:
        st.markdown("###  Top Performers")
        col_2nd, col_1st, col_3rd = st.columns([1, 1, 1])
        
        with col_1st:
            st.markdown("""
            <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%); 
                        border-radius: 15px; margin: 10px;">
                <div style="font-size: 48px;">🥇</div>
                <div style="font-size: 24px; font-weight: bold; color: white;">1st Place</div>
                <div style="font-size: 20px; color: white; margin-top: 10px;">{}</div>
                <div style="font-size: 16px; color: white;">Score: {:.1f}</div>
            </div>
            """.format(rankings[0]['name'], rankings[0]['score']), unsafe_allow_html=True)
        
        with col_2nd:
            st.markdown("""
            <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #C0C0C0 0%, #A8A8A8 100%); 
                        border-radius: 15px; margin: 10px; margin-top: 40px;">
                <div style="font-size: 40px;">🥈</div>
                <div style="font-size: 20px; font-weight: bold; color: white;">2nd Place</div>
                <div style="font-size: 18px; color: white; margin-top: 10px;">{}</div>
                <div style="font-size: 14px; color: white;">Score: {:.1f}</div>
            </div>
            """.format(rankings[1]['name'], rankings[1]['score']), unsafe_allow_html=True)
        
        with col_3rd:
            st.markdown("""
            <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #CD7F32 0%, #B8860B 100%); 
                        border-radius: 15px; margin: 10px; margin-top: 60px;">
                <div style="font-size: 36px;">🥉</div>
                <div style="font-size: 18px; font-weight: bold; color: white;">3rd Place</div>
                <div style="font-size: 16px; color: white; margin-top: 10px;">{}</div>
                <div style="font-size: 14px; color: white;">Score: {:.1f}</div>
            </div>
            """.format(rankings[2]['name'], rankings[2]['score']), unsafe_allow_html=True)
        
        st.markdown("---")
    
    # Full Rankings Table
    st.markdown("###  Complete Rankings")
    
    # Create DataFrame for display
    ranking_data = []
    for r in rankings:
        # Get historical average
        avg_ranking = get_average_ranking(manager_id, r['employee_id'], months=6)
        
        # Determine trend
        if avg_ranking and avg_ranking['months_tracked'] > 1:
            if r['rank'] < avg_ranking['avg_rank']:
                trend = "📈 Improving"
                trend_color = "#10b981"
            elif r['rank'] > avg_ranking['avg_rank']:
                trend = "📉 Declining"
                trend_color = "#ef4444"
            else:
                trend = "➡️ Stable"
                trend_color = "#64748b"
        else:
            trend = "🆕 New"
            trend_color = "#3b82f6"
        
        ranking_data.append({
            'Rank': r['rank'],
            'Name': r['name'],
            'Designation': r['designation'],
            'Total Goals': r['total_goals'],
            'Completed': r['completed_goals'],
            'Completion %': f"{r['completion_rate']:.1f}%",
            'Avg Progress %': f"{r['avg_progress']:.1f}%",
            'Score': f"{r['score']:.1f}",
            'Avg Rank (6m)': f"{avg_ranking['avg_rank']:.1f}" if avg_ranking else "N/A",
            'Trend': trend,
            'employee_id': r['employee_id']
        })
    
    df_rankings = pd.DataFrame(ranking_data)
    
    # Display with styling
    st.dataframe(
        df_rankings.drop('employee_id', axis=1),
        use_container_width=True,
        height=400
    )
    
    # Ranking Insights
    st.markdown("---")
    st.markdown("### 💡 Ranking Insights")
    
    col_insight1, col_insight2, col_insight3 = st.columns(3)
    
    with col_insight1:
        top_performer = rankings[0]
        st.success(f"""
        **🏆 Top Performer**
        
        {top_performer['name']}
        
        - Rank: #{top_performer['rank']}
        - Score: {top_performer['score']:.1f}
        - Completion: {top_performer['completion_rate']:.1f}%
        """)
    
    with col_insight2:
        avg_score = sum(r['score'] for r in rankings) / len(rankings)
        st.info(f"""
        **📊 Team Average**
        
        - Avg Score: {avg_score:.1f}
        - Avg Completion: {sum(r['completion_rate'] for r in rankings) / len(rankings):.1f}%
        - Team Size: {len(rankings)}
        """)
    
    with col_insight3:
        needs_support = [r for r in rankings if r['score'] < 50]
        st.warning(f"""
        **⚠️ Needs Support**
        
        {len(needs_support)} team member(s) scoring below 50
        
        {'- ' + needs_support[0]['name'] if needs_support else 'All team members performing well! ✅'}
        """)
    
    # Individual Performance History
    st.markdown("---")
    st.markdown("### Individual Performance History")
    
    selected_member = st.selectbox(
        "Select Team Member",
        [r['name'] for r in rankings],
        key="history_member"
    )
    
    if selected_member:
        member_data = next(r for r in rankings if r['name'] == selected_member)
        employee_id = member_data['employee_id']
        
        col_hist1, col_hist2 = st.columns([2, 1])
        
        with col_hist1:
            # Get historical data
            history = get_historical_rankings(manager_id, employee_id, months=12)
            
            if history:
                # Create trend chart
                history_df = pd.DataFrame([
                    {
                        'Month': f"{get_month_name(h['month'])} {h['year']}",
                        'Rank': h['rank'],
                        'Score': h['score'],
                        'Completion %': h['completion_rate']
                    }
                    for h in reversed(history)
                ])
                
                # Rank trend (lower is better, so invert for chart)
                fig = make_subplots(
                    rows=2, cols=1,
                    subplot_titles=('Ranking Trend', 'Performance Score Trend'),
                    vertical_spacing=0.15
                )
                
                fig.add_trace(
                    go.Scatter(
                        x=history_df['Month'],
                        y=history_df['Rank'],
                        mode='lines+markers',
                        name='Rank',
                        line=dict(color='#3b82f6', width=3),
                        marker=dict(size=10)
                    ),
                    row=1, col=1
                )
                
                fig.add_trace(
                    go.Scatter(
                        x=history_df['Month'],
                        y=history_df['Score'],
                        mode='lines+markers',
                        name='Score',
                        line=dict(color='#10b981', width=3),
                        marker=dict(size=10)
                    ),
                    row=2, col=1
                )
                
                fig.update_xaxes(title_text="Month", row=2, col=1)
                fig.update_yaxes(title_text="Rank (Lower is Better)", row=1, col=1, autorange="reversed")
                fig.update_yaxes(title_text="Score", row=2, col=1)
                
                fig.update_layout(height=600, showlegend=False)
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No historical data available yet. Save monthly rankings to track progress over time.")
        
        with col_hist2:
            avg_stats = get_average_ranking(manager_id, employee_id, months=6)
            
            if avg_stats:
                st.markdown("**6-Month Statistics**")
                st.metric("Average Rank", f"#{avg_stats['avg_rank']:.1f}")
                st.metric("Best Rank", f"#{avg_stats['best_rank']}")
                st.metric("Worst Rank", f"#{avg_stats['worst_rank']}")
                st.metric("Months Tracked", avg_stats['months_tracked'])
                
                # Performance indicator
                current_rank = member_data['rank']
                if current_rank < avg_stats['avg_rank']:
                    st.success("📈 Performing above average!")
                elif current_rank > avg_stats['avg_rank']:
                    st.warning("📉 Below recent average")
                else:
                    st.info("➡️ Consistent performance")
            else:
                st.info("Need at least 2 months of data for statistics")

def render_password_strength_meter(password, key_suffix=""):
    """Render password strength meter with visual feedback"""
    score, color, strength, feedback = check_password_strength(password)
    
    # Strength meter HTML
    st.markdown(f"""
    <div style="margin: 15px 0;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <span style="font-size: 13px; font-weight: 600; color: #64748b;">Password Strength:</span>
            <span style="font-size: 14px; font-weight: bold; color: {color};">{strength}</span>
        </div>
        <div style="width: 100%; height: 8px; background: #e2e8f0; border-radius: 10px; overflow: hidden;">
            <div style="width: {score}%; height: 100%; background: {color}; transition: width 0.3s ease;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Show feedback if password is not very strong
    if score < 85 and feedback:
        with st.expander(" Tips to strengthen your password", expanded=False):
            for tip in feedback:
                st.markdown(f"- {tip}")
            
            st.markdown("""
            **Best Practices:**
            - Use a mix of uppercase and lowercase letters
            - Include numbers and special characters
            - Avoid common words or patterns
            - Use at least 8-12 characters
            - Don't reuse passwords from other accounts
            """)
def send_password_reset_email(email, reset_token):
    """Send password reset email"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            st.error("Email configuration not found. Please contact administrator.")
            return False
        
        # Create reset link (you'll need to handle this in your app)
        reset_link = f"http://your-app-url.com/reset-password?token={reset_token}"
        
        # Email content
        subject = "Password Reset Request - Performance Management System"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #3b82f6; text-align: center;">🎯 Password Reset Request</h2>
                    <p>Hello,</p>
                    <p>You requested to reset your password for the Performance Management System.</p>
                    <p><strong>Your Reset Token:</strong></p>
                    <div style="background: #f3f4f6; padding: 15px; border-radius: 5px; text-align: center; font-size: 24px; font-weight: bold; letter-spacing: 2px; color: #1e40af;">
                        {reset_token}
                    </div>
                    <p style="margin-top: 20px;">Please use this token to reset your password. This token will expire in 1 hour.</p>
                    <p style="color: #ef4444;"><strong>⚠️ Important:</strong> If you did not request this password reset, please ignore this email or contact your administrator.</p>
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="font-size: 12px; color: #64748b; text-align: center;">
                        This is an automated email from Performance Management System.<br>
                        Please do not reply to this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        # Create message
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, email, message.as_string())
        
        return True
    except Exception as e:
        st.error(f"Failed to send email: {str(e)}")
        return False


# PAGE CONFIG MUST BE FIRST STREAMLIT COMMAND
st.set_page_config(
    page_title="Performance Management System",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Load environment variables
load_dotenv()

# Initialize Supabase client ONCE
supabase = get_supabase_client()

from helper import apply_theme
apply_theme()

# Import our modules
from database import Database
from helper import (
    init_session_state, get_quarter_months, get_month_name,
    get_quarter_name, calculate_progress, format_goal_table_data,
    render_user_avatar, render_card, render_metric_card, render_progress_bar,
    render_feedback_card, validate_goal_data
)
# Initialize session state and database
init_session_state()
db = Database()
# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

# ============================================
# SESSION PERSISTENCE
# ============================================
def save_session_to_storage():
    """Save session state to browser storage"""
    if st.session_state.user:
        try:
            # Store in Streamlit's query params (URL-based persistence)
            st.query_params['user_id'] = str(st.session_state.user['id'])
            st.query_params['page'] = st.session_state.page
            
            # ✅ Save navigation state
            if st.session_state.get('selected_year'):
                st.query_params['year'] = str(st.session_state.selected_year)
            if st.session_state.get('selected_quarter'):
                st.query_params['quarter'] = str(st.session_state.selected_quarter)
            if st.session_state.get('selected_month'):
                st.query_params['month'] = str(st.session_state.selected_month)
            
            # ✅ Save active tab states
            if st.session_state.get('active_month_tab'):
                st.query_params['month_tab'] = str(st.session_state.active_month_tab)
            if st.session_state.get('active_hr_tab'):
                st.query_params['hr_tab'] = str(st.session_state.active_hr_tab)
            
            # ✅ Save employee viewing state
            if st.session_state.get('viewing_employee'):
                st.query_params['viewing_emp_id'] = str(st.session_state.viewing_employee['id'])
            
            if st.session_state.get('viewing_employee_year'):
                st.query_params['viewing_emp_year'] = 'true'
            
        except Exception as e:
            print(f"Session save error: {str(e)}")


def restore_session_from_storage():
    """Restore session state from browser storage"""
    try:
        # Check URL params for user_id
        user_id = st.query_params.get('user_id')
        if user_id:
            # Restore user from database
            user = db.get_user_by_id(user_id)
            if user:
                st.session_state.user = user
                st.session_state.page = st.query_params.get('page', 'dashboard')
                
                # ✅ Restore navigation state
                year_param = st.query_params.get('year')
                if year_param:
                    st.session_state.selected_year = int(year_param)
                
                quarter_param = st.query_params.get('quarter')
                if quarter_param:
                    st.session_state.selected_quarter = int(quarter_param)
                
                month_param = st.query_params.get('month')
                if month_param:
                    st.session_state.selected_month = int(month_param)
                
                # ✅ Restore tab states
                month_tab = st.query_params.get('month_tab')
                if month_tab:
                    st.session_state.active_month_tab = int(month_tab)
                
                hr_tab = st.query_params.get('hr_tab')
                if hr_tab:
                    st.session_state.active_hr_tab = int(hr_tab)
                
                # ✅ Restore employee viewing state
                viewing_emp_id = st.query_params.get('viewing_emp_id')
                if viewing_emp_id:
                    viewing_employee = db.get_user_by_id(viewing_emp_id)
                    if viewing_employee:
                        st.session_state.viewing_employee = viewing_employee
                
                viewing_emp_year = st.query_params.get('viewing_emp_year')
                if viewing_emp_year == 'true':
                    st.session_state.viewing_employee_year = True
                
                return True
    except Exception as e:
        print(f"Session restore error: {str(e)}")
    
    return False


# ✅ RESTORE SESSION ON PAGE LOAD (if not already logged in)
if not st.session_state.user:
    restore_session_from_storage()

# ✅ SAVE SESSION ON CHANGES (if user is logged in)
if st.session_state.user:
    save_session_to_storage()



# ============================================
# WEEK DATE CALCULATION HELPERS
# ============================================
def get_week_dates(year, month, week_num):
    """Get start and end dates for a specific week in a month"""
    # Get first day of month
    first_day = date(year, month, 1)
    # Get last day of month
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    
    # Calculate week ranges (approximate 7-day periods)
    if week_num == 1:
        start_date = first_day
        end_date = min(first_day + timedelta(days=6), last_day)
    elif week_num == 2:
        start_date = first_day + timedelta(days=7)
        end_date = min(first_day + timedelta(days=13), last_day)
    elif week_num == 3:
        start_date = first_day + timedelta(days=14)
        end_date = min(first_day + timedelta(days=20), last_day)
    else:  # week 4
        start_date = first_day + timedelta(days=21)
        end_date = last_day
    
    return start_date, end_date


def get_week_for_date(year, month, target_date):
    """Determine which week a date falls into"""
    first_day = date(year, month, 1)
    day_diff = (target_date - first_day).days
    
    if day_diff < 7:
        return 1
    elif day_diff < 14:
        return 2
    elif day_diff < 21:
        return 3
    else:
        return 4

# ============================================
# LOGIN PAGE
# ============================================
def login_page():
    """Display login page with forgot password option"""
    st.markdown("<h1 style='text-align: center;'> Performance Management System</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #64748b;'>Sign in to continue</p>", unsafe_allow_html=True)
    
    # Check if showing forgot password or reset password
    if 'show_forgot_password' not in st.session_state:
        st.session_state.show_forgot_password = False
    
    if 'show_reset_password' not in st.session_state:
        st.session_state.show_reset_password = False
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # ===== NORMAL LOGIN FORM =====
        if not st.session_state.show_forgot_password and not st.session_state.show_reset_password:
            with st.form("login_form"):
                email = st.text_input(" Email", placeholder="your@email.com")
                password = st.text_input("Password", type="password", placeholder="••••••••")
                submit = st.form_submit_button("Sign In", use_container_width=True)
                
                if submit:
                    if email and password:
                        user = db.authenticate_user(email, password)
                        if user:
                            st.session_state.user = user
                            st.session_state.page = 'dashboard'  # ✅ Set default page
                            save_session_to_storage()  # ✅ Save session
                            st.success("✅ Login successful!")
                            st.rerun()
                        else:
                            st.error(" Invalid credentials")
                    else:
                        st.warning(" Please enter both email and password")
            
            # Forgot Password Link
            col_fp1, col_fp2, col_fp3 = st.columns([1, 2, 1])
            with col_fp2:
                if st.button(" Forgot Password?", use_container_width=True):
                    st.session_state.show_forgot_password = True
                    st.rerun()
            
            
        
        # ===== FORGOT PASSWORD FORM =====
        elif st.session_state.show_forgot_password:
            st.markdown("###  Forgot Password")
            st.info("Enter your email address and we'll send you a reset token.")
            
            with st.form("forgot_password_form"):
                reset_email = st.text_input("Email Address", placeholder="your@email.com")
                submit_reset = st.form_submit_button("Send Reset Token", use_container_width=True)
                
                if submit_reset:
                    if reset_email:
                        # Generate token
                        token = db.create_password_reset_token(reset_email)
                        
                        if token:
                            # Send email
                            if send_password_reset_email(reset_email, token):
                                st.success("✅ Password reset token sent to your email!")
                                st.info(f"**Your Reset Token:** `{token}`\n\nPlease check your email for the reset token. It will expire in 1 hour.")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                            else:
                                st.warning(f"⚠️ Could not send email, but here's your reset token: **{token}**")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                        else:
                            st.error("❌ Email not found in system")
                    else:
                        st.warning("⚠️ Please enter your email address")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_forgot_password = False
                st.rerun()
        
        # ===== RESET PASSWORD FORM =====
        elif st.session_state.show_reset_password:
            st.markdown("### Reset Password")
            st.info("Enter the reset token you received via email and your new password.")
            
            with st.form("reset_password_form"):
                reset_token = st.text_input("Reset Token", placeholder="Enter 8-character token")
                new_password = st.text_input(" New Password", type="password", placeholder="••••••••", key="login_new_pass")
                
                # Password strength meter
                if new_password:
                    render_password_strength_meter(new_password, "login_reset")
                
                confirm_password = st.text_input("Confirm Password", type="password", placeholder="••••••••", key="login_confirm_pass")
                submit_new_password = st.form_submit_button("Reset Password", use_container_width=True)
                
                if submit_new_password:
                    if reset_token and new_password and confirm_password:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if score < 30:
                                    st.warning(f" Your password is {strength}. Consider making it stronger for better security.")
                                
                                # Reset password
                                if db.reset_password_with_token(reset_token, new_password):
                                    st.success("✅ Password reset successful! You can now login with your new password.")
                                    st.balloons()
                                    st.session_state.show_reset_password = False
                                    st.query_params.clear()
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_reset_password = False
                st.rerun()

# ============================================
# ENHANCED DASHBOARD PAGE
# ============================================
def display_dashboard():
    """Display dashboard with enhanced charts"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    st.title(" Dashboard")
    if 'display_auto_complete_banner' in globals():
        display_auto_complete_banner()
    # User Stats Section - Now in Dashboard
    st.subheader(" Your Performance")
    stats = db.get_user_goal_stats(user['id'])
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    with col_stat1:
        render_metric_card("Total Goals", str(stats.get('total_goals', 0)), color="#3b82f6")
    with col_stat2:
        render_metric_card("Completed", str(stats.get('completed_goals', 0)), color="#10b981")
    with col_stat3:
        render_metric_card("Active", str(stats.get('active_goals', 0)), color="#f59e0b")
    with col_stat4:
        render_metric_card("Avg Progress", f"{stats.get('avg_progress', 0):.1f}%", color="#8b5cf6")
    
    st.markdown("---")
    
    # Get data based on role
    if role == 'HR':
        all_users = db.get_all_users()
        total_employees = len([u for u in all_users if u['role'] == 'Employee'])
        total_managers = len([u for u in all_users if u['role'] == 'Manager'])
        
        st.subheader(" Organization Overview")
        
        # Stats Row
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            render_metric_card("Total Employees", str(total_employees), color="#3b82f6")
        with col2:
            render_metric_card("Total Managers", str(total_managers), color="#8b5cf6")
        with col3:
            render_metric_card("Total Users", str(len(all_users)), color="#10b981")
        with col4:
            active_goals = len(db.get_all_active_goals())
            render_metric_card("Active Goals", str(active_goals), color="#f59e0b")
        
        st.markdown("---")
        
        # Employee Rankings
        st.subheader(" Employee Performance Rankings")
        rankings = []
        for emp in [u for u in all_users if u['role'] == 'Employee']:
            emp_stats = db.get_user_goal_stats(emp['id'])
            if emp_stats.get('total_goals', 0) > 0:
                rankings.append({
                    'Name': emp['name'],
                    'Department': emp.get('department', 'N/A'),
                    'Total Goals': emp_stats.get('total_goals', 0),
                    'Completed': emp_stats.get('completed_goals', 0),
                    'Progress %': f"{emp_stats.get('avg_progress', 0):.1f}%",
                    'Progress_Val': emp_stats.get('avg_progress', 0)
                })

        if rankings:
            df_rank = pd.DataFrame(rankings)
            df_rank = df_rank.sort_values('Progress_Val', ascending=False)
            df_rank = df_rank.drop('Progress_Val', axis=1)
            df_rank.insert(0, 'Rank', range(1, len(df_rank) + 1))
            
            # 🔥 Hide the dataframe index
            st.dataframe(df_rank, use_container_width=True, hide_index=True)

        
    elif role == 'Manager':
        team_members = db.get_team_members(user['id'])
        
        st.subheader(" Team Overview")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            render_metric_card("Team Members", str(len(team_members)), color="#3b82f6")
        with col2:
            my_stats = db.get_user_goal_stats(user['id'])
            render_metric_card("My Goals", str(my_stats.get('total_goals', 0)), color="#8b5cf6")
        with col3:
            team_goals = sum([len(db.get_user_all_goals(m['id'])) for m in team_members])
            render_metric_card("Team Goals", str(team_goals), color="#10b981")
        st.markdown("---")
        display_team_rankings_dashboard(user['id'])

        st.markdown("---")
        st.subheader(" Team Performance")
        
        team_perf = []
        for member in team_members:
            member_stats = db.get_user_goal_stats(member['id'])
            team_perf.append({
                'Name': member['name'],
                'Designation': member.get('designation', 'N/A'),
                'Total Goals': member_stats.get('total_goals', 0),
                'Completed': member_stats.get('completed_goals', 0),
                'Progress %': f"{member_stats.get('avg_progress', 0):.1f}%"
            })
        
        if team_perf:
            df_team = pd.DataFrame(team_perf)
            st.dataframe(df_team, use_container_width=True)
        else:
            st.info("No team members assigned yet")
    
    # Enhanced Notifications with Week Deadlines
    st.markdown("---")
    st.subheader(" Reminders & Notifications")

    notifications = get_enhanced_notifications(user)
    if notifications:
        for notif in notifications:
            # Create a nice card layout for each notification
            notif_type = notif['type']
            
            if notif_type == 'assignment':
                icon = "📬"
                color = "#3b82f6"
            elif notif_type == 'deadline':
                icon = "⏰"
                color = "#f59e0b"
            elif notif_type == 'overdue':
                icon = "🚨"
                color = "#ef4444"
            elif notif_type == 'team_incomplete':
                icon = "⚠️"
                color = "#f97316"
            elif notif_type == 'feedback':
                icon = "💬"
                color = "#8b5cf6"
            elif notif_type == 'achievement':
                icon = "🎉"
                color = "#10b981"
            elif notif_type == 'team_overdue':
                icon = "👥"
                color = "#ef4444"
            else:
                icon = "ℹ️"
                color = "#64748b"
            
            # Display notification with time
            st.markdown(f'''
            <div style="background: white; padding: 12px 16px; border-radius: 8px; border-left: 4px solid {color}; 
                        margin-bottom: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                <div style="display: flex; justify-content: space-between; align-items: start;">
                    <div style="flex: 1;">
                        <div style="font-weight: 600; color: #1e293b; margin-bottom: 4px;">
                            {icon} {notif['title']}
                        </div>
                        <div style="color: #475569; font-size: 14px;">
                            {notif['message']}
                        </div>
                    </div>
                    <div style="font-size: 12px; color: #94a3b8; white-space: nowrap; margin-left: 16px;">
                        {notif.get('time', '')}
                    </div>
                </div>
            </div>
            ''', unsafe_allow_html=True)
    else:
        st.info("No new notifications")


def get_enhanced_notifications(user):
    """Get enhanced notifications including feedback notifications"""
    notifications = []
    role = user['role']
    today = date.today()
    now = datetime.now(IST)
    
    # Get user's goals
    goals = db.get_user_all_goals(user['id'])
    
    # ✅ NEW: Check for new feedback first (highest priority)
    for goal in goals:
        goal_feedback = db.get_goal_feedback(goal['goal_id'])
        for feedback in goal_feedback:
            # Only show feedback from others (not self-feedback)
            if feedback.get('feedback_by') != user['id']:
                created_at = feedback.get('created_at')
                if created_at:
                    try:
                        if isinstance(created_at, str):
                            if 'T' in created_at:
                                fb_datetime = datetime.strptime(created_at[:19], '%Y-%m-%dT%H:%M:%S')
                                fb_datetime = pytz.utc.localize(fb_datetime).astimezone(IST)
                            else:
                                fb_datetime = datetime.strptime(created_at[:10], '%Y-%m-%d')
                                fb_datetime = IST.localize(fb_datetime)
                        else:
                            fb_datetime = created_at
                            if fb_datetime.tzinfo is None:
                                fb_datetime = IST.localize(fb_datetime)
                        
                        fb_date = fb_datetime.date()
                        days_since = (today - fb_date).days
                        
                        # Show feedback from last 7 days
                        if 0 <= days_since <= 7:
                            time_str = fb_datetime.strftime('%I:%M %p')
                            date_str = fb_datetime.strftime('%b %d, %Y')
                            
                            if days_since == 0:
                                time_ago = f"Today at {time_str}"
                            elif days_since == 1:
                                time_ago = f"Yesterday at {time_str}"
                            else:
                                time_ago = f"{date_str} at {time_str}"
                            
                            feedback_type = feedback.get('feedback_type', 'Feedback')
                            rating = feedback.get('rating', 0)
                            rating_stars = '⭐' * rating
                            
                            notifications.append({
                                'type': 'feedback',
                                'title': f'New {feedback_type} Received',
                                'message': f"{feedback.get('feedback_by_name', 'Someone')} rated '{goal['goal_title']}' {rating_stars} - {feedback.get('comment', '')[:50]}...",
                                'time': time_ago,
                                'timestamp': fb_datetime,
                                'priority': 0,  # Highest priority
                                'created_order': len(notifications)
                            })
                    except Exception as e:
                        pass
    
    # ✅ Check for newly assigned goals (priority 1)
    for goal in goals:
        created_at = goal.get('created_at')
        if created_at:
            try:
                if isinstance(created_at, str):
                    if 'T' in created_at:
                        created_datetime = datetime.strptime(created_at[:19], '%Y-%m-%dT%H:%M:%S')
                        created_datetime = pytz.utc.localize(created_datetime).astimezone(IST)
                    else:
                        created_datetime = datetime.strptime(created_at[:10], '%Y-%m-%d')
                        created_datetime = IST.localize(created_datetime)
                else:
                    created_datetime = created_at
                    if created_datetime.tzinfo is None:
                        created_datetime = IST.localize(created_datetime)
                
                created_date = created_datetime.date()
                days_since_creation = (today - created_date).days
                
                if 0 <= days_since_creation <= 7:
                    assigner_id = goal.get('created_by')
                    if assigner_id and assigner_id != user['id']:
                        assigner = db.get_user_by_id(assigner_id)
                        if assigner:
                            time_str = created_datetime.strftime('%I:%M %p')
                            date_str = created_datetime.strftime('%b %d, %Y')
                            
                            if days_since_creation == 0:
                                time_ago = f"Today at {time_str}"
                            elif days_since_creation == 1:
                                time_ago = f"Yesterday at {time_str}"
                            else:
                                time_ago = f"{date_str} at {time_str}"
                            
                            notifications.append({
                                'type': 'assignment',
                                'title': 'New Goal Assigned',
                                'message': f"{assigner['name']} assigned you goal '{goal['goal_title']}'",
                                'time': time_ago,
                                'timestamp': created_datetime,
                                'priority': 1,
                                'created_order': len(notifications)
                            })
            except Exception as e:
                pass
    
    # Check for achievements (priority 3)
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        if progress >= 100 and goal.get('status') == 'Active':
            end_date_str = goal.get('end_date')
            try:
                if end_date_str:
                    achievement_datetime = datetime.strptime(str(end_date_str), '%Y-%m-%d')
                else:
                    achievement_datetime = datetime.now()
                
                achievement_datetime = IST.localize(achievement_datetime.replace(hour=12, minute=0))
                achievement_date = achievement_datetime.date()
                
                days_ago = (today - achievement_date).days
                if days_ago == 0:
                    time_ago = "Today"
                elif days_ago == 1:
                    time_ago = "Yesterday"
                else:
                    time_ago = achievement_datetime.strftime('%b %d, %Y')
                
                notifications.append({
                    'type': 'achievement',
                    'title': 'Goal Completed!',
                    'message': f"Congratulations! You completed '{goal['goal_title']}'",
                    'time': time_ago,
                    'timestamp': achievement_datetime,
                    'priority': 3,
                    'created_order': len(notifications)
                })
            except:
                pass
    
    # Check weekly deadlines and overdue items (priority 1-2)
    for goal in goals:
        if goal.get('status') == 'Active':
            end_date_str = goal.get('end_date')
            if end_date_str:
                try:
                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                    days_left = (end_date - today).days
                    
                    end_datetime = IST.localize(datetime.combine(end_date, datetime.min.time()).replace(hour=23, minute=59))
                    
                    if days_left < 0:
                        notifications.append({
                            'type': 'overdue',
                            'title': 'Goal Overdue',
                            'message': f"Goal '{goal['goal_title']}' is overdue by {abs(days_left)} days",
                            'time': end_date.strftime('%b %d, %Y'),
                            'timestamp': end_datetime,
                            'priority': 0,
                            'created_order': len(notifications)
                        })
                    elif 0 <= days_left <= 3:
                        if days_left == 0:
                            time_msg = "Due today"
                        elif days_left == 1:
                            time_msg = "Due tomorrow"
                        else:
                            time_msg = f"Due in {days_left} days"
                        
                        notifications.append({
                            'type': 'deadline',
                            'title': 'Goal Deadline Approaching',
                            'message': f"Goal '{goal['goal_title']}' - {time_msg}",
                            'time': end_date.strftime('%b %d, %Y'),
                            'timestamp': end_datetime,
                            'priority': 1,
                            'created_order': len(notifications)
                        })
                except:
                    pass
    
    # Manager notifications for incomplete team member goals
    if role == 'Manager':
        team_members = db.get_team_members(user['id'])
        for member in team_members:
            member_goals = db.get_user_all_goals(member['id'])
            
            for goal in member_goals:
                if goal.get('status') == 'Active':
                    end_date_str = goal.get('end_date')
                    if end_date_str:
                        try:
                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                            end_datetime = IST.localize(datetime.combine(end_date, datetime.min.time()).replace(hour=23, minute=59))
                            
                            if today > end_date:
                                monthly_achievement = goal.get('monthly_achievement', 0)
                                monthly_target = goal.get('monthly_target', 1)
                                progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                                
                                if progress < 100:
                                    notifications.append({
                                        'type': 'team_incomplete',
                                        'title': 'Team Goal Incomplete',
                                        'message': f"{member['name']} did not complete '{goal['goal_title']}' (Progress: {progress:.1f}%)",
                                        'time': end_date.strftime('%b %d, %Y'),
                                        'timestamp': end_datetime,
                                        'priority': 2,
                                        'created_order': len(notifications)
                                    })
                        except:
                            pass
    
    # Sort notifications: Priority first, then most recent timestamp
    notifications.sort(key=lambda x: (
        x.get('priority', 999),  # Lower priority number = higher importance
        -(x.get('timestamp', datetime.min.replace(tzinfo=IST)).timestamp())  # Most recent first
    ))
    
    return notifications[:15]

# ============================================
# VIEW ALL GOALS PAGE (NEW)
# ============================================
def display_view_all_goals():
    """View all goals of a user with edit/delete options"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("Session expired. Please login again.")
        st.rerun()

    st.title(" View All Goals")
    
    # Select user (HR can see all, Manager can see team, Employee sees own)
    if role == 'HR':
        all_users = db.get_all_users()
        selected_user = st.selectbox(
            "Select User",
            [f"{u['name']} ({u['email']})" for u in all_users]
        )
        user_email = selected_user.split('(')[1].strip(')')
        selected_user_obj = next(u for u in all_users if u['email'] == user_email)
        view_user_id = selected_user_obj['id']
    elif role == 'Manager':
        team_members = db.get_team_members(user['id'])
        if team_members:
            selected_user = st.selectbox(
                "Select Team Member",
                [user['name']] + [f"{m['name']} ({m['email']})" for m in team_members]
            )
            if selected_user == user['name']:
                view_user_id = user['id']
            else:
                user_email = selected_user.split('(')[1].strip(')')
                selected_user_obj = next(m for m in team_members if m['email'] == user_email)
                view_user_id = selected_user_obj['id']
        else:
            view_user_id = user['id']
    else:
        view_user_id = user['id']
    
    # Get all goals for selected user
    all_goals = db.get_user_all_goals(view_user_id)
    
    if not all_goals:
        st.info("No goals found for this user")
        return
    
    # Filter options
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_year = st.selectbox("Filter by Year", ["All"] + sorted(list(set([g['year'] for g in all_goals])), reverse=True))
    with col2:
        filter_status = st.selectbox("Filter by Status", ["All", "Active", "Completed", "On Hold", "Cancelled"])
    with col3:
        search_term = st.text_input("🔍 Search Goal Title")
    
    # Apply filters
    filtered_goals = all_goals
    if filter_year != "All":
        filtered_goals = [g for g in filtered_goals if g['year'] == filter_year]
    if filter_status != "All":
        filtered_goals = [g for g in filtered_goals if g.get('status') == filter_status]
    if search_term:
        filtered_goals = [g for g in filtered_goals if search_term.lower() in g['goal_title'].lower()]
    
    st.markdown(f"**Showing {len(filtered_goals)} of {len(all_goals)} goals**")
    
    # Display goals in expandable cards
    for goal in filtered_goals:
        progress = calculate_progress(goal.get('monthly_achievement', 0), goal.get('monthly_target', 1))
        
        # Determine status color
        if progress >= 100:
            status_color = "#10b981"
            status_text = "Completed"
        elif progress >= 60:
            status_color = "#f59e0b"
            status_text = "On Track"
        else:
            status_color = "#ef4444"
            status_text = "At Risk"
        
        with st.expander(f"🎯 {goal['goal_title']} - {goal['year']}/Q{goal.get('quarter', 'N/A')}/M{goal.get('month', 'N/A')}"):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"**Vertical:** {goal.get('vertical', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
            
            with col2:
                st.markdown(f"**Start:** {goal['start_date']}")
                st.markdown(f"**End:** {goal['end_date']}")
            
            with col3:
                st.markdown(f"**Target:** {goal.get('monthly_target', 0)}")
                st.markdown(f"**Achievement:** {goal.get('monthly_achievement', 0)}")
            
            with col4:
                st.markdown(f"**Status:** <span style='color: {status_color}; font-weight: bold;'>{goal.get('status', 'Active')}</span>", unsafe_allow_html=True)
                st.markdown(f"**Progress:** <span style='color: {status_color}; font-weight: bold;'>{progress:.1f}%</span>", unsafe_allow_html=True)
            
            st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
            
            # Edit and Delete buttons
            col_edit, col_delete, col_space = st.columns([1, 1, 3])
            
            with col_edit:
                if st.button(" Edit", key=f"edit_goal_{goal['goal_id']}", use_container_width=True):
                    st.session_state.editing_goal = goal
                    st.rerun()
            
            with col_delete:
                if st.button(" Delete", key=f"delete_goal_{goal['goal_id']}", use_container_width=True):
                    if db.delete_goal(goal['goal_id']):
                        st.success("Goal deleted!")
                        st.rerun()
    
    # Edit Goal Modal
    if 'editing_goal' in st.session_state:
        st.markdown("---")
        st.subheader(" Edit Goal")
        
        edit_goal = st.session_state.editing_goal
        
        with st.form("edit_all_goals_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_vertical = st.text_input("Vertical*", value=edit_goal.get('vertical', ''))
                new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                new_monthly_target = st.number_input("Monthly Target", min_value=0.0, value=float(edit_goal.get('monthly_target', 0)))
            
            with col2:
                new_status = st.selectbox("Status", ['Active', 'Completed', 'On Hold', 'Cancelled'],
                                         index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active')))
                new_monthly_achievement = st.number_input("Monthly Achievement", min_value=0.0, value=float(edit_goal.get('monthly_achievement', 0)))
                new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button(" Save Changes", use_container_width=True):
                    updates = {
                        'vertical': new_vertical,
                        'goal_title': new_title,
                        'kpi': new_kpi,
                        'monthly_target': new_monthly_target,
                        'monthly_achievement': new_monthly_achievement,
                        'goal_description': new_description,
                        'status': new_status
                    }
                    if db.update_goal(edit_goal['goal_id'], updates):
                        st.success("✅ Goal updated!")
                        del st.session_state.editing_goal
                        st.rerun()
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_goal
                    st.rerun()



# ============================================
# HR INFO PAGE (FIXED)
# ============================================
def display_hr_info():
    """Display all HR information with delete user option"""
    user = st.session_state.user
    
    if not user:
        st.warning(" Session expired. Please login again.")
        st.rerun()

    if user['role'] != 'HR':
        st.warning(" Only HR can access this page")
        return
    
    st.title(" HR Information Dashboard")
    
    all_users = db.get_all_users()
    
    # HR Stats (keep existing)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        total_users = len(all_users)
        render_metric_card("Total Users", str(total_users), color="#3b82f6")
    with col2:
        total_goals = sum([len(db.get_user_all_goals(u['id'])) for u in all_users])
        render_metric_card("Total Goals", str(total_goals), color="#8b5cf6")
    with col3:
        total_feedback = len(db.get_all_feedback())
        render_metric_card("Total Feedback", str(total_feedback), color="#10b981")
    with col4:
        active_goals = len(db.get_all_active_goals())
        render_metric_card("Active Goals", str(active_goals), color="#f59e0b")
    
    st.markdown("---")
    
    # Tabs for different views
    tab1, tab2, tab3, tab4 = st.tabs(["👥 All Users", "📊 Department Stats", "🎯 Goal Summary", "💬 Feedback Summary"])
    
    with tab1:
        st.subheader("All Users in System")
        
        users_data = []
        for u in all_users:
            stats = db.get_user_goal_stats(u['id'])
            manager_name = "N/A"
            if u.get('manager_id'):
                manager = db.get_user_by_id(u['manager_id'])
                if manager:
                    manager_name = manager['name']
            
            users_data.append({
                'Name': u['name'],
                'Email': u['email'],
                'Role': u['role'],
                'Department': u.get('department', 'N/A'),
                'Manager': manager_name,
                'Total Goals': stats.get('total_goals', 0),
                'Completed': stats.get('completed_goals', 0),
                'Progress %': f"{stats.get('avg_progress', 0):.1f}%",
                'User ID': u['id']
            })
        
        df_users = pd.DataFrame(users_data)
        
        # Display without User ID column
        display_df = df_users.drop('User ID', axis=1)
        st.dataframe(display_df, use_container_width=True, height=500)
        
        # Delete User option - can delete ALL users except current logged in user
        st.markdown("---")
        st.subheader("⚠️ Delete User (Admin Only)")

        # Show all users except the current logged-in one
        deletable_users = [u for u in all_users if u['id'] != user['id']]

        if deletable_users:
            col_del1, col_del2 = st.columns([2, 1])
            with col_del1:
                delete_user_select = st.selectbox(
                    "Select User to Delete",
                    [f"{u['name']} ({u['role']}) - {u['email']}" for u in deletable_users]
                )

            with col_del2:
                if st.button("Delete Selected User", type="primary", use_container_width=True):
                    # Save selected user in session state (so Streamlit remembers after rerun)
                    user_email = delete_user_select.split(' - ')[1]
                    delete_user_obj = next(u for u in deletable_users if u['email'] == user_email)
                    st.session_state['user_to_delete'] = delete_user_obj

        # Show confirmation prompt if a user is selected
        if 'user_to_delete' in st.session_state:
            delete_user_obj = st.session_state['user_to_delete']

            st.warning(f"Are you sure you want to delete **{delete_user_obj['name']}** ({delete_user_obj['role']})?")
            st.error("⚠️ This will also delete all their goals and feedback!")

            confirm_col1, confirm_col2 = st.columns(2)
            with confirm_col1:
                if st.button("Yes, Delete", key="confirm_delete_user"):
                    result = db.delete_user(delete_user_obj['id'])
                    if result:
                        st.success(f"✅ User {delete_user_obj['name']} deleted successfully!")
                        del st.session_state['user_to_delete']
                        st.rerun()
                    else:
                        st.error("❌ Failed to delete user. Check Supabase connection or permissions.")

            with confirm_col2:
                if st.button("Cancel", key="cancel_delete_user"):
                    del st.session_state['user_to_delete']
                    st.info("❎ Deletion cancelled.")
        else:
            st.info("No other users available for deletion")

    
    with tab2:
        st.subheader("Department-wise Statistics")
        
        dept_stats = {}
        for u in all_users:
            dept = u.get('department', 'Unassigned')
            if dept not in dept_stats:
                dept_stats[dept] = {'users': 0, 'goals': 0, 'completed': 0}
            dept_stats[dept]['users'] += 1
            stats = db.get_user_goal_stats(u['id'])
            dept_stats[dept]['goals'] += stats.get('total_goals', 0)
            dept_stats[dept]['completed'] += stats.get('completed_goals', 0)
        
        dept_data = []
        for dept, stats in dept_stats.items():
            dept_data.append({
                'Department': dept,
                'Total Users': stats['users'],
                'Total Goals': stats['goals'],
                'Completed Goals': stats['completed'],
                'Completion Rate': f"{(stats['completed'] / stats['goals'] * 100) if stats['goals'] > 0 else 0:.1f}%"
            })
        
        df_dept = pd.DataFrame(dept_data)
        st.dataframe(df_dept, use_container_width=True)
    
    with tab3:
        st.subheader("Goal Summary")
        
        all_goals = []
        for u in all_users:
            all_goals.extend(db.get_user_all_goals(u['id']))
        
        # Status breakdown
        status_count = {}
        for goal in all_goals:
            status = goal.get('status', 'Active')
            status_count[status] = status_count.get(status, 0) + 1
        
        col_status1, col_status2, col_status3, col_status4 = st.columns(4)
        with col_status1:
            st.metric("Active", status_count.get('Active', 0))
        with col_status2:
            st.metric("Completed", status_count.get('Completed', 0))
        with col_status3:
            st.metric("On Hold", status_count.get('On Hold', 0))
        with col_status4:
            st.metric("Cancelled", status_count.get('Cancelled', 0))
        
        # Year-wise breakdown
        st.markdown("---")
        st.subheader("Year-wise Goal Distribution")
        
        year_stats = {}
        for goal in all_goals:
            year = goal['year']
            if year not in year_stats:
                year_stats[year] = 0
            year_stats[year] += 1
        
        if year_stats:
            # Create dataframe for proper plotting
            df_year = pd.DataFrame({
                'Year': list(year_stats.keys()),
                'Goals': list(year_stats.values())
            })
            
            fig = px.bar(
                df_year,
                x='Year',
                y='Goals',
                title='Goals by Year',
                color='Goals',
                color_continuous_scale='Blues'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No goals data available")
    
    with tab4:
        st.subheader("Feedback Summary")
        
        all_feedback = db.get_all_feedback()
        
        # Feedback type breakdown
        type_count = {}
        for fb in all_feedback:
            fb_type = fb.get('feedback_type', 'Unknown')
            type_count[fb_type] = type_count.get(fb_type, 0) + 1
        
        col_fb1, col_fb2, col_fb3 = st.columns(3)
        with col_fb1:
            st.metric("Manager Feedback", type_count.get('Manager', 0))
        with col_fb2:
            st.metric("HR Feedback", type_count.get('HR', 0))
        with col_fb3:
            st.metric("Self Appraisals", type_count.get('Self Appraisal', 0))
        
        # Average ratings
        st.markdown("---")
        st.subheader("Average Ratings")
        
        if all_feedback:
            total_rating = sum([fb.get('rating', 0) for fb in all_feedback])
            avg_rating = total_rating / len(all_feedback) if all_feedback else 0
            
            st.metric("Overall Average Rating", f"{avg_rating:.2f} ⭐")
            
            # Rating distribution
            st.markdown("---")
            st.subheader("Rating Distribution")
            
            rating_count = {}
            for fb in all_feedback:
                rating = fb.get('rating', 0)
                rating_count[rating] = rating_count.get(rating, 0) + 1
            
            df_rating = pd.DataFrame({
                'Rating': list(rating_count.keys()),
                'Count': list(rating_count.values())
            })
            
            fig2 = px.bar(
                df_rating,
                x='Rating',
                y='Count',
                title='Feedback Rating Distribution',
                color='Rating',
                color_continuous_scale='RdYlGn'
            )
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No feedback data available")


# ============================================
# EMPLOYEES PAGE
# ============================================
def display_employees_page():
    """Display all employees with cards and assign goal option"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    st.title("👥 Employees")
    
    # Get employees based on role
    if role == 'HR':
        employees = [u for u in db.get_all_users() if u['role'] in ['Employee', 'Manager']]
    elif role == 'Manager':
        employees = db.get_team_members(user['id'])
    else:
        st.warning("⚠️ You don't have permission to view this page")
        return
    
    if not employees:
        st.info("No employees found")
        return
    
    # Assign Goal Section for HR and Manager
    with st.expander("➕ Assign Goal to Employee"):
        display_quick_assign_goal_form(user, employees)
    
    st.markdown("---")
    
    # Search and Filter
    col1, col2 = st.columns([2, 1])
    with col1:
        search = st.text_input("Search by name or email", "")
    with col2:
        filter_dept = st.selectbox("Filter by Department", ["All"] + list(set([e.get('department', 'N/A') for e in employees])))
    
    # Filter employees
    filtered_employees = employees
    if search:
        filtered_employees = [e for e in filtered_employees if search.lower() in e['name'].lower() or search.lower() in e['email'].lower()]
    if filter_dept != "All":
        filtered_employees = [e for e in filtered_employees if e.get('department') == filter_dept]
    
    # Display employee cards
    st.markdown("---")
    cols = st.columns(3)
    for idx, emp in enumerate(filtered_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px;'>
                            {emp['role']}
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            col_view, col_edit, col_del = st.columns(3)
            
            with col_view:
                if st.button("👁️", key=f"view_emp_{emp['id']}", use_container_width=True, help="View Goals"):
                    st.session_state.viewing_employee = emp
                    st.session_state.page = 'employee_goals'
                    st.rerun()
            
            # Edit and Delete for HR only
            if role in ['HR', 'Manager']:
                with col_edit:
                    if st.button("✏️", key=f"edit_emp_{emp['id']}", use_container_width=True, help="Edit Employee"):
                        st.session_state.editing_employee = emp
                        st.rerun()
                
                with col_del:
                    if st.button("🗑️", key=f"del_emp_{emp['id']}", use_container_width=True, help="Delete Employee"):
                        st.session_state.deleting_employee = emp
                        st.rerun()
    
    # Edit Employee Modal - Show at top with expander
    if 'editing_employee' in st.session_state:
        st.markdown("---")
        
        
        edit_emp = st.session_state.editing_employee
        
        # Prominent header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">✏️ Editing: {edit_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("edit_employee_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*", value=edit_emp['name'])
                new_email = st.text_input("Email*", value=edit_emp['email'])
                new_designation = st.text_input("Designation", value=edit_emp.get('designation', ''))
            
            with col2:
                new_role = st.selectbox(
                    "Role*", 
                    ["Employee", "Manager", "HR"],
                    index=["Employee", "Manager", "HR"].index(edit_emp['role'])
                )
                new_department = st.text_input("Department", value=edit_emp.get('department', ''))
                
                # Manager assignment
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager' and u['id'] != edit_emp['id']]
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                
                current_manager_idx = 0
                if edit_emp.get('manager_id'):
                    current_manager = db.get_user_by_id(edit_emp['manager_id'])
                    if current_manager:
                        current_manager_str = f"{current_manager['name']} ({current_manager['email']})"
                        if current_manager_str in manager_options:
                            current_manager_idx = manager_options.index(current_manager_str)
                
                selected_manager = st.selectbox("Assign to Manager", manager_options, index=current_manager_idx)
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button("Save Changes", use_container_width=True):
                    if new_name and new_email and new_role:
                        manager_id = None
                        if selected_manager != "None":
                            manager_email = selected_manager.split('(')[1].strip(')')
                            manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                        
                        updates = {
                            'name': new_name,
                            'email': new_email.lower().strip(),
                            'designation': new_designation,
                            'role': new_role,
                            'department': new_department,
                            'manager_id': manager_id
                        }
                        
                        if db.update_user(edit_emp['id'], updates):
                            st.success(f"✅ Employee {new_name} updated successfully!")
                            del st.session_state.editing_employee
                            st.rerun()
                        else:
                            st.error("❌ Failed to update employee")
                    else:
                        st.error("❌ Please fill all required fields")
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_employee
                    st.rerun()
    
    # Delete Employee Modal - Show at top with expander
    if 'deleting_employee' in st.session_state:
        st.markdown("---")
        st.markdown("---")
        
        del_emp = st.session_state.deleting_employee
        
        # Prominent warning header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">⚠️ Delete Employee: {del_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.warning(f"**Are you sure you want to delete {del_emp['name']}?**")
        st.error("⚠️ This will also delete all their goals and feedback! This action cannot be undone.")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info(f"**Name:** {del_emp['name']}")
        with col2:
            st.info(f"**Email:** {del_emp['email']}")
        with col3:
            st.info(f"**Role:** {del_emp['role']}")
        
        # Get stats
        emp_goals = db.get_user_all_goals(del_emp['id'])
        st.warning(f"This will delete **{len(emp_goals)} goals** associated with this employee")
        
        confirm = st.checkbox("I understand this action cannot be undone", key="confirm_emp_delete")
        
        col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
        
        with col_del2:
            if st.button(
                "🗑️ Delete Employee", 
                disabled=not confirm,
                use_container_width=True,
                type="primary",
                key="execute_emp_delete"
            ):
                if db.delete_user(del_emp['id']):
                    st.success(f"✅ Employee {del_emp['name']} deleted successfully!")
                    del st.session_state.deleting_employee
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Failed to delete employee")
        
        with col_del3:
            if st.button("❌ Cancel", use_container_width=True, key="cancel_emp_delete"):
                del st.session_state.deleting_employee
                st.rerun()


def display_quick_assign_goal_form(user, employees):
    """Quick assign goal form in employees page"""
    with st.form("quick_assign_goal"):
        st.subheader("Quick Assign Goal")
        
        col1, col2 = st.columns(2)
        
        with col1:
            selected_emp = st.selectbox(
                "Assign To*",
                [f"{e['name']} ({e['email']})" for e in employees]
            )
            
            emp_email = selected_emp.split('(')[1].strip(')')
            emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
            
            year = st.number_input("Year", min_value=2020, max_value=2100, value=datetime.now().year)
            quarter = st.selectbox("Quarter", [1, 2, 3, 4])
            month = st.selectbox("Month", list(range(1, 13)), index=datetime.now().month - 1)
        
        with col2:
            vertical = st.text_input("Vertical*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        # Auto-fill dates based on month
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            start_date = st.date_input("Start Date", value=month_start)
        with col_date2:
            end_date = st.date_input("End Date", value=month_end)
        
        description = st.text_area("Description")
        monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        st.markdown("**Weekly Targets (Auto-calculated or Manual)**")
        col3, col4, col5, col6 = st.columns(4)
        
        auto_divide = st.checkbox("Auto-divide monthly target into weeks")
        
        if auto_divide:
            week_target = monthly_target / 4
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, value=week_target, key="quick_w1")
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, value=week_target, key="quick_w2")
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, value=week_target, key="quick_w3")
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, value=week_target, key="quick_w4")
        else:
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, key="quick_w1_manual")
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, key="quick_w2_manual")
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, key="quick_w3_manual")
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, key="quick_w4_manual")
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Remarks (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_remarks = st.text_area("Week 1 Remarks", key="quick_w1_remarks", height=80)
        with col_r2:
            w2_remarks = st.text_area("Week 2 Remarks", key="quick_w2_remarks", height=80)
        with col_r3:
            w3_remarks = st.text_area("Week 3 Remarks", key="quick_w3_remarks", height=80)
        with col_r4:
            w4_remarks = st.text_area("Week 4 Remarks", key="quick_w4_remarks", height=80)
        # ========== END OF ADDED SECTION ==========
        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if vertical and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'vertical': vertical,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_remarks,  # ✅ ADD
                    'week2_remarks': w2_remarks,  # ✅ ADD
                    'week3_remarks': w3_remarks,  # ✅ ADD
                    'week4_remarks': w4_remarks,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    st.success(f"✅ Goal assigned to {selected_emp}")
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")



# ============================================
# EMPLOYEE GOALS VIEW
# ============================================
def display_employee_goals():
    """Display goals for a specific employee"""
    if not st.session_state.get('viewing_employee'):
        st.warning("⚠️ Employee data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    emp = st.session_state.viewing_employee
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back to Employees"):
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"📊 {emp['name']}'s Goals")
    
    # Show year selection for this employee
    years = db.get_years(emp['id'])
    
    if not years:
        st.info(f"No goals found for {emp['name']}")
        return
    
    st.subheader("📆 Select Year")
    
    # Display years in rows with goal counts
    sorted_years = sorted(years.items(), reverse=True)
    for year, summary in sorted_years:
        # Get goal count for this year
        year_goals = [g for g in db.get_user_all_goals(emp['id']) if g['year'] == year]
        goal_count = len(year_goals)
        
        st.markdown(f"""
        <div class='hierarchy-card' style='cursor: pointer;'>
            <h2 style='margin:0;'>📅 {year} <span style='color: #64748b; font-size: 16px;'>({goal_count} goals)</span></h2>
            <p style='color: #64748b; margin-top: 8px;'>{summary[:80] if summary else 'Click to view quarters'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button(f"View {year}", key=f"emp_year_{year}", use_container_width=True):
            st.session_state.selected_year = year
            st.session_state.viewing_employee_year = True
            st.session_state.page = 'employee_quarters'
            st.rerun()
        
        st.markdown("<br>", unsafe_allow_html=True)


# ============================================
# MY GOALS PAGE (UPDATED)
# ============================================
def display_my_goals():
    """Display user's own goals with month search"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.session_state.page = 'login'
        st.rerun()
    
    st.title(f" My Goals - {user['name']}")
    st.caption(f"{user.get('designation', 'Employee')} • {user['role']}")
    
    # Month Quick Search
    st.markdown("---")
    st.subheader("🔍 Quick Search by Month")
    
    col_search1, col_search2 = st.columns([2, 1])
    with col_search1:
        search_month = st.selectbox(
            "Select Month to View Across All Years",
            ["None"] + [get_month_name(i) for i in range(1, 13)]
        )
    
    if search_month != "None":
        month_num = [get_month_name(i) for i in range(1, 13)].index(search_month) + 1
        
        st.subheader(f" {search_month} Goals Across All Years")
        
        all_goals = db.get_user_all_goals(user['id'])
        month_goals = [g for g in all_goals if g.get('month') == month_num]
        
        if month_goals:
            # Group by year
            year_groups = {}
            for goal in month_goals:
                year = goal['year']
                if year not in year_groups:
                    year_groups[year] = []
                year_groups[year].append(goal)
            
            for year in sorted(year_groups.keys(), reverse=True):
                with st.expander(f"{search_month} {year} ({len(year_groups[year])} goals)"):
                    for goal in year_groups[year]:
                        progress = calculate_progress(
                            goal.get('monthly_achievement', 0),
                            goal.get('monthly_target', 1)
                        )
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col1:
                            st.markdown(f"**{goal['goal_title']}**")
                            st.caption(f"Vertical: {goal.get('vertical', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                        with col2:
                            st.metric("Target", goal.get('monthly_target', 0))
                        with col3:
                            st.metric("Progress", f"{progress:.1f}%")
                        
                        if st.button(f"View Goal", key=f"view_month_goal_{goal['goal_id']}"):
                            st.session_state.selected_year = year
                            st.session_state.selected_quarter = goal.get('quarter')
                            st.session_state.selected_month = month_num
                            st.session_state.page = 'month_goals'
                            st.rerun()
                        
                        st.markdown("---")
        else:
            st.info(f"No goals found for {search_month}")
        
        st.markdown("---")
    
    # Regular year display
    # Regular year display
    years = db.get_years(user['id'])
    current_year = datetime.now().year
    if current_year not in years:
        years[current_year] = ""
    
    st.subheader(" Browse by Year")
    
    # Add Create New Year button
    col_header1, col_header2 = st.columns([3, 1])
    with col_header2:
        if st.button("➕ Create New Year", use_container_width=True, key="create_new_year_btn"):
            st.session_state.creating_new_year = True
    
    # Show create year form if button clicked
    if st.session_state.get('creating_new_year'):
        with st.expander("➕ Create New Year", expanded=True):
            with st.form("create_new_year_form"):
                new_year = st.number_input("Year*", min_value=2020, max_value=2100, value=datetime.now().year + 1)
                new_year_summary = st.text_area("Year Summary (Optional)", placeholder="Enter goals/plans for this year...")
                
                col_create1, col_create2 = st.columns(2)
                with col_create1:
                    if st.form_submit_button("✅ Create Year", use_container_width=True):
                        if new_year:
                            # Check if year already exists
                            if new_year in years:
                                st.error(f"❌ Year {new_year} already exists!")
                            else:
                                # Create year entry
                                if db.update_year_summary(user['id'], new_year, new_year_summary):
                                    st.success(f"✅ Year {new_year} created successfully!")
                                    st.session_state.creating_new_year = False
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to create year")
                        else:
                            st.error("❌ Please enter a valid year")
                
                with col_create2:
                    if st.form_submit_button("❌ Cancel", use_container_width=True):
                        st.session_state.creating_new_year = False
                        st.rerun()
        st.markdown("---")
    
    # Display years in rows with goal counts
    sorted_years = sorted(years.items(), reverse=True)
    for year, summary in sorted_years:
        # Get goal count for this year
        year_goals = [g for g in db.get_user_all_goals(user['id']) if g['year'] == year]
        goal_count = len(year_goals)
        
        st.markdown(f"""
        <div class='hierarchy-card' style='cursor: pointer;'>
            <h2 style='margin:0;'> {year} <span style='color: #64748b; font-size: 16px;'>({goal_count} goals)</span></h2>
            <p style='color: #64748b; margin-top: 8px;'>{summary[:80] if summary else 'Click to view quarters'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 1])
        
        with col_btn1:
            if st.button(f"View {year}", key=f"year_view_{year}", use_container_width=True):
                st.session_state.selected_year = year
                st.session_state.page = 'quarters'
                st.rerun()
        
        with col_btn2:
            if st.button("✏️ Edit", key=f"year_edit_{year}", use_container_width=True):
                st.session_state.editing_year = year
                st.session_state.editing_year_summary = summary
        
        with col_btn3:
            if st.button("🗑️ Delete", key=f"year_del_{year}", use_container_width=True):
                if db.delete_year(user['id'], year):
                    st.success(f"Year {year} deleted!")
                    st.rerun()
        
        st.markdown("<br>", unsafe_allow_html=True)
    
    # Rest of the existing code for edit year and add new year
    # ...

# ============================================
# QUARTER SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_quarter_selection():
    """Display quarter selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    
    if not year:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    # Check if viewing employee goals
    # Check if viewing employee goals
    if st.session_state.get('viewing_employee_year'):
        # ✅ Safety check for employee
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back"):
                st.session_state.viewing_employee_year = False
                st.session_state.page = 'employee_goals'
                st.rerun()
        with col2:
            st.title(f"{emp['name']}'s Year {year} - Quarters")
        user_id = emp['id']
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Years"):
                st.session_state.page = 'my_goals'
                st.rerun()
        with col2:
            st.title(f"📊 Year {year} - Quarters")
        user_id = user['id']
    
    quarters = db.get_quarters(user_id, year)
    for q in [1, 2, 3, 4]:
        if q not in quarters:
            quarters[q] = ""
    
    cols = st.columns(2)
    for idx, (quarter, summary) in enumerate(sorted(quarters.items())):
        with cols[idx % 2]:
            # Get goal count for this quarter
            quarter_goals = [g for g in db.get_user_all_goals(user_id) if g['year'] == year and g.get('quarter') == quarter]
            goal_count = len(quarter_goals)
            
            st.markdown(f"<div class='q-btn-wrapper-{quarter}'>", unsafe_allow_html=True)
            st.markdown(f"""
            <div class='hierarchy-card'>
                <h2>📈 Quarter {quarter} <span style='color: #64748b; font-size: 14px;'>({goal_count} goals)</span></h2>
                <p style='color: #64748b;'>{get_quarter_name(quarter)}</p>
                <p style='margin-top: 8px;'>{summary[:80] if summary else 'Click to view months'}</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button(f"Open Q{quarter}", key=f"q_{quarter}", use_container_width=True,  type="primary" ):
                st.session_state.selected_quarter = quarter
                st.session_state.page = 'employee_months' if st.session_state.get('viewing_employee_year') else 'months'
                st.rerun()
            
            # 🔹 Close wrapper
            st.markdown("</div>", unsafe_allow_html=True)
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit Q{quarter} Summary"):
                    with st.form(f"edit_q{quarter}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"qsum_{quarter}")
                        if st.form_submit_button("Save"):
                            if db.update_quarter_summary(user_id, year, quarter, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_month_selection():
    """Display month selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    
    if not year or not quarter:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    
    # Check if viewing employee goals
    if st.session_state.get('viewing_employee_year'):
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Quarters"):
                st.session_state.page = 'employee_quarters'
                st.rerun()
        with col2:
            st.title(f" {emp['name']}'s Year {year} - Q{quarter} - Months")
        user_id = emp['id']
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Quarters"):
                st.session_state.page = 'quarters'
                st.rerun()
        with col2:
            st.title(f" Year {year} - Q{quarter} - Months")
        user_id = user['id']
    
    quarter_month_nums = get_quarter_months(quarter)
    months = db.get_months(user_id, year, quarter)
    
    cols = st.columns(3)
    for idx, month_num in enumerate(quarter_month_nums):
        with cols[idx]:
            month_name = get_month_name(month_num)
            summary = months.get(month_num, "")
            month_goals = db.get_month_goals(user_id, year, quarter, month_num)
            goal_count = len(month_goals)

            st.markdown(f"""
            <div class='month-card'>
                <div class='month-card-content'>
                    <h2> {month_name}</h2>
                    <p>{goal_count} goals</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button(
                f"Open {month_name}",
                key=f"m_{user_id}_{year}_{quarter}_{month_num}", 
                use_container_width=True):
                st.session_state.selected_month = month_num
                st.session_state.page = 'employee_month_goals' if st.session_state.get('viewing_employee_year') else 'month_goals'
                st.rerun()
            
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit {month_name} Summary"):
                    with st.form(f"edit_m{month_num}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"msum_{month_num}")
                        if st.form_submit_button("Save"):
                            if db.update_month_summary(user_id, year, quarter, month_num, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH GOALS PAGE (WITH ASSIGN IN MONTHLY VIEW)
# ============================================
def display_month_goals():
    """Display month goals with week tabs"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    month = st.session_state.selected_month

    if not year or not quarter or not month:
        st.warning("⚠️ Navigation state lost. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()

    month_name = get_month_name(month)
    
    if st.session_state.get('viewing_employee_year'):
        # ✅ Safety check for employee
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Months"):
                st.session_state.page = 'employee_months'
                st.rerun()
        with col2:
            st.title(f" {emp['name']}'s {month_name} {year} Goals")
        display_user = emp
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Months"):
                st.session_state.page = 'months'
                st.rerun()
        with col2:
            st.title(f"{month_name} {year} Goal Sheet")
        display_user = user
    
    # Create tabs
    tab_all, tab_w1, tab_w2, tab_w3, tab_w4 = st.tabs([
        "📋 Monthly View",
        "📅 Week 1",
        "📅 Week 2",
        "📅 Week 3",
        "📅 Week 4"
    ])
    
    # Monthly view
    with tab_all:
        display_monthly_view(display_user, year, quarter, month)
    
    # Week views
    for week_num, tab in enumerate([tab_w1, tab_w2, tab_w3, tab_w4], 1):
        with tab:
            display_week_view(display_user, year, quarter, month, week_num)

def export_goals_to_excel(user_id, year, quarter, month):
    """Export goals to Excel with proper formatting including Monthly Achievement"""
    
    goals = db.get_month_goals(user_id, year, quarter, month)
    
    if not goals:
        st.warning("No goals to export")
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{get_month_name(month)} {year}"
    
    # Define styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Vertical", "Goal Title", "KPI", "Monthly Target", "Start Date", "End Date"
    ]
    
    # Weekly headers
    weekly_headers = ["Week 1", "Week 2", "Week 3", "Week 4"]
    
    # Row 1: Main headers
    current_col = 1
    for header in headers:
        cell = ws.cell(row=1, column=current_col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Add "Weekly Target" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Target"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Weekly Achievement" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Achievement"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Monthly Achievement" header (single column)
    cell = ws.cell(row=1, column=current_col, value="Monthly Achievement")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    
    # Add "Remarks" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Remarks"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    
    # Row 2: Week subheaders
    current_col = 7  # Start after main headers
    
    # Week numbers under "Weekly Target"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Week numbers under "Weekly Achievement"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Skip Monthly Achievement column (already merged)
    current_col += 1
    
    # Week numbers under "Remarks"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Fill main header cells in row 2
    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.border = border
    
    # Data rows
    row_num = 3
    for goal in goals:
        col_num = 1
        
        # Main goal info
        ws.cell(row=row_num, column=col_num, value=goal.get('vertical', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal['goal_title']).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('kpi', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('monthly_target', 0)).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('start_date', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('end_date', '')).border = border
        col_num += 1
        
        # Weekly targets
        for week in range(1, 5):
            ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_target', 0)).border = border
            col_num += 1
        
        # Weekly achievements
        for week in range(1, 5):
            ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_achievement', 0)).border = border
            col_num += 1
        
        # Monthly achievement
        cell = ws.cell(row=row_num, column=col_num, value=goal.get('monthly_achievement', 0))
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        col_num += 1
        
        # Weekly remarks
        for week in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_remarks', ''))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_num += 1
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Vertical
    ws.column_dimensions['B'].width = 30  # Goal Title
    ws.column_dimensions['C'].width = 15  # KPI
    ws.column_dimensions['D'].width = 15  # Monthly Target
    ws.column_dimensions['E'].width = 12  # Start Date
    ws.column_dimensions['F'].width = 12  # End Date
    
    # Weekly columns (targets and achievements)
    for col in range(7, 15):  # Weeks 1-4 for targets and achievements
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Monthly Achievement column
    ws.column_dimensions['O'].width = 15
    
    # Remarks columns (wider)
    for col in range(16, 20):
        ws.column_dimensions[chr(64 + col)].width = 25
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
# ============================================
# MONTHLY VIEW (WITH ASSIGN GOAL)
# ============================================
def display_monthly_view(user, year, quarter, month):
    """Display monthly goals view with Excel-like format and assign goal option"""
    if not st.session_state.get('viewing_employee_year'):
        col_title, col_button = st.columns([3, 1])
        with col_title:
            st.subheader("📊 Monthly Goals Overview")
        
        
        # Show create goal form if button clicked
        if st.session_state.get('show_create_goal_form'):
            with st.expander("➕ Create New Monthly Goal", expanded=True):
                display_add_goal_form_inline(user, year, quarter, month)
            st.markdown("---")
    else:
        st.subheader("📊 Monthly Goals Overview")
    
    # Assign Goal Section (for HR and Manager) - NOW IN MONTHLY VIEW
    if user['role'] in ['HR', 'Manager'] and not st.session_state.get('viewing_employee_year'):
        with st.expander("➕ Assign Goal to Employee"):
            display_assign_goal_form_monthly(user, year, quarter, month)
    
    
    # Get goals
    goals = db.get_month_goals(user['id'], year, quarter, month)
    
    # Add export button
    # Create & Export Buttons - side by side
    col1, col2 = st.columns([1, 1], gap="small")

    with col1:
        if st.button("➕ Create New Goal", key="create_goal_month", use_container_width=True):
            st.session_state.show_create_goal_form = True

    with col2:
        if st.button("📥 Export Excel", key="export_excel_month", use_container_width=True):
            excel_file = export_goals_to_excel(user['id'], year, quarter, month)
            if excel_file:
                st.download_button(
                    label="📄 Download Excel",
                    data=excel_file,
                    file_name=f"Goals_{get_month_name(month)}_{year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


    st.markdown("---")
    if goals:
        # Display goals in Excel-like format
        st.markdown("### 📋 Monthly Goal Sheet")
        
        # Create the formatted table data
        table_data = []
        for goal in goals:
            row = {
                'Vertical': goal.get('vertical', ''),
                'Goal Title': goal['goal_title'],
                'KPI': goal.get('kpi', ''),
                'Monthly Target': goal.get('monthly_target', 0),
                'Start Date': goal.get('start_date', ''),
                'End Date': goal.get('end_date', ''),
                # Weekly Targets
                'Week 1 Target': goal.get('week1_target', 0),
                'Week 2 Target': goal.get('week2_target', 0),
                'Week 3 Target': goal.get('week3_target', 0),
                'Week 4 Target': goal.get('week4_target', 0),
                # Weekly Achievements
                'Week 1 Achievement': goal.get('week1_achievement', 0),
                'Week 2 Achievement': goal.get('week2_achievement', 0),
                'Week 3 Achievement': goal.get('week3_achievement', 0),
                'Week 4 Achievement': goal.get('week4_achievement', 0),
                # Monthly Achievement (sum of weekly)
                'Monthly Achievement': goal.get('monthly_achievement', 0),
                # Weekly Remarks
                'Week 1 Remarks': goal.get('week1_remarks', ''),
                'Week 2 Remarks': goal.get('week2_remarks', ''),
                'Week 3 Remarks': goal.get('week3_remarks', ''),
                'Week 4 Remarks': goal.get('week4_remarks', '')
            }
            table_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(table_data)
        
        # Display with better styling using custom HTML/CSS
        st.markdown("""
<style>

.excel-table {
    width: 100%;
    overflow-x: auto;
    margin-top: 10px;
}

/* Table Base */
.excel-table table {
    border-collapse: collapse;
    width: 100%;
    font-size: 13px;
    background: #ffffff;
    color: #1e293b;
    border-radius: 10px;
    overflow: hidden;
    border: 1px solid #e2e8f0;
    box-shadow: 0 4px 12px rgba(0,0,0,0.04);
}

/* Header */
.excel-table th,
.excel-table .section-header {
    background: linear-gradient(180deg, #3b82f6, #2563eb);
    color: white;
    padding: 12px;
    text-align: center;
    border: 1px solid #2563eb;
    font-weight: 600;
    letter-spacing: 0.5px;
    position: sticky;
    top: 0;
    z-index: 10;
    text-transform: uppercase;
    font-size: 12px;
}

/* Table Cells */
.excel-table td {
    border: 1px solid #e5e7eb;
    padding: 10px;
    text-align: center;
    background-color: #ffffff;
    vertical-align: middle;
    transition: background 0.2s ease;
}

/* Target columns */
.excel-table .target-col {
    background-color: #eef5ff;
}

/* Achievement columns */
.excel-table .achievement-col {
    background-color: #edfdf4;
}

/* Monthly Achievement */
.excel-table .monthly-achievement-col {
    background-color: #dffbe7;
    font-weight: 700;
    color: #166534;
}

/* Remarks columns */
.excel-table .remarks-col {
    background-color: #fffceb;
    color: #444;
    word-wrap: break-word;
    max-width: 220px;
    line-height: 1.4;
}

/* Date Columns */
.excel-table td:nth-child(5),
.excel-table td:nth-child(6) {
    text-align: center;
    white-space: nowrap;
    font-size: 12px;
    color: #475569;
}

/* Hover Row */
.excel-table tr:hover td {
    background-color: #f8fafc;
}

/* Bold key columns */
.excel-table th,
.excel-table td:first-child,
.excel-table td:nth-child(2) {
    font-weight: 600;
}

/* Goal title button */
.goal-title-btn {
    background: none;
    border: none;
    color: #2563eb;
    text-decoration: underline;
    font-weight: 600;
    cursor: pointer;
    padding: 0;
    font-size: inherit;
    transition: color 0.2s ease;
}

.goal-title-btn:hover {
    color: #1e40af;
}

/* Modal styling */
.goal-modal {
    display: none;
    position: fixed;
    z-index: 10000;
    left: 0;
    top: 0;
    width: 100%;
    height: 100%;
    background-color: rgba(15,23,42,0.55);
    backdrop-filter: blur(4px);
    overflow: auto;
}

.modal-content {
    background-color: #ffffff;
    margin: 6% auto;
    padding: 26px;
    border-radius: 12px;
    width: 80%;
    max-width: 650px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.25);
    position: relative;
    animation: slideIn 0.3s ease-out;
}

@keyframes slideIn {
    from { transform: translateY(-40px); opacity: 0; }
    to { transform: translateY(0); opacity: 1; }
}

.close-btn {
    position: absolute;
    right: 18px;
    top: 18px;
    background: #ef4444;
    color: white;
    border: none;
    border-radius: 6px;
    padding: 6px 12px;
    cursor: pointer;
    font-weight: 600;
    font-size: 14px;
    transition: background 0.2s ease;
}

.close-btn:hover {
    background: #dc2626;
}

/* Responsive */
@media (max-width: 768px) {
    .excel-table table {
        font-size: 11px;
    }
    .excel-table th,
    .excel-table td {
        padding: 6px;
    }
}

</style>
""", unsafe_allow_html=True)


        # Store the complete HTML in a variable for components
        complete_html = f'''
        <!DOCTYPE html>
        <html>
        <head>
        <style>
        .excel-table {{
            width: 100%;
            overflow-x: auto;
        }}

        .excel-table table {{
            border-collapse: collapse;
            width: 100%;
            font-size: 13px;
            background-color: #ffffff;
            color: #222222;
            border: 1px solid #ddd;
        }}

        .excel-table th {{
            background-color: #3B82F6;
            color: #ffffff;
            padding: 10px;
            text-align: center;
            border: 1px solid #cfcfcf;
            font-weight: 600;
        }}

        .excel-table td {{
            border: 1px solid #e0e0e0;
            padding: 8px;
            text-align: center;
            background-color: #ffffff;
            vertical-align: middle;
        }}

        .target-col {{ background-color: #F3F8FF; }}
        .achievement-col {{ background-color: #F6FFF8; }}
        .monthly-achievement-col {{ background-color: #E9FCEB; font-weight: bold; }}
        .remarks-col {{ background-color: #FFFDF4; max-width: 180px; word-wrap: break-word; }}

        .goal-title-btn {{
            background: none;
            border: none;
            color: #3B82F6;
            text-decoration: underline;
            font-weight: bold;
            cursor: pointer;
            padding: 0;
            font-size: inherit;
        }}

        .goal-title-btn:hover {{
            color: #1E40AF;
        }}

        .modal {{
            display: none;
            position: fixed;
            z-index: 10000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow: auto;
        }}

        .modal-content {{
            background-color: #fefefe;
            margin: 5% auto;
            padding: 20px;
            border-radius: 10px;
            width: 80%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            position: relative;
            animation: slideIn 0.3s ease-out;
        }}

        @keyframes slideIn {{
            from {{ transform: translateY(-50px); opacity: 0; }}
            to {{ transform: translateY(0); opacity: 1; }}
        }}

        .close-btn {{
            position: absolute;
            right: 15px;
            top: 15px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 15px;
            cursor: pointer;
            font-weight: bold;
            font-size: 18px;
        }}

        .close-btn:hover {{
            background: #dc2626;
        }}
        </style>
        </head>
        <body>

        <div class="excel-table">
        <table>
        <thead>
        <tr>
            <th rowspan="2">Vertical</th>
            <th rowspan="2">Goal Title</th>
            <th rowspan="2">KPI</th>
            <th rowspan="2">Monthly Target</th>
            <th rowspan="2">Start Date</th>
            <th rowspan="2">End Date</th>
            <th colspan="4">Weekly Target</th>
            <th colspan="4">Weekly Achievement</th>
            <th rowspan="2">Monthly Achievement</th>
            <th colspan="4">Remarks</th>
        </tr>
        <tr>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
        </tr>
        </thead>
        <tbody>
        '''

        # Add data rows
        for goal_idx, (goal, (_, row)) in enumerate(zip(goals, df.iterrows())):
            goal_desc = goal.get('goal_description', 'No description available').replace("'", "\\'").replace('"', '\\"').replace('\n', '<br>')
            
            complete_html += f'''
        <tr>
            <td>{row["Vertical"]}</td>
            <td>
                <button class="goal-title-btn" onclick="openModal({goal_idx})">
                    {row["Goal Title"]}
                </button>
            </td>
            <td>{row["KPI"]}</td>
            <td>{row["Monthly Target"]}</td>
            <td>{row["Start Date"]}</td>
            <td>{row["End Date"]}</td>
        '''
            
            # Weekly targets
            for week in range(1, 5):
                complete_html += f'<td class="target-col">{row[f"Week {week} Target"]}</td>'
            
            # Weekly achievements
            for week in range(1, 5):
                achievement = row[f"Week {week} Achievement"]
                target = row[f"Week {week} Target"]
                progress = (achievement / target * 100) if target > 0 else 0
                color = '#4CAF50' if progress >= 100 else '#FFC107' if progress >= 60 else '#F44336'
                complete_html += f'<td class="achievement-col" style="color: {color}; font-weight: bold;">{achievement}</td>'
            
            # Monthly achievement
            monthly_achievement = row["Monthly Achievement"]
            monthly_target = row["Monthly Target"]
            monthly_progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
            monthly_color = '#4CAF50' if monthly_progress >= 100 else '#FFC107' if monthly_progress >= 60 else '#F44336'
            complete_html += f'<td class="monthly-achievement-col" style="color: {monthly_color};">{monthly_achievement}</td>'
            
            # Weekly remarks
            for week in range(1, 5):
                remarks = row[f"Week {week} Remarks"]
                complete_html += f'<td class="remarks-col">{remarks if remarks else "-"}</td>'
            
            complete_html += '</tr>'
            
            # Add modal for this goal
            complete_html += f'''
        <div id="modal{goal_idx}" class="modal">
            <div class="modal-content">
                <button class="close-btn" onclick="closeModal({goal_idx})">✕</button>
                <h3 style="margin: 0 0 15px 0; color: #1E3A8A; padding-right: 40px;">Goal Description</h3>
                <div style="color: #333; line-height: 1.6; max-height: 400px; overflow-y: auto; 
                    padding: 15px; background: #f9fafb; border-radius: 5px;">
                    {goal_desc}
                </div>
            </div>
        </div>
        '''

        complete_html += '''
        </tbody>
        </table>
        </div>

        <script>
        function openModal(index) {
            document.getElementById('modal' + index).style.display = 'block';
        }

        function closeModal(index) {
            document.getElementById('modal' + index).style.display = 'none';
        }

        // Close modal when clicking outside
        window.onclick = function(event) {
            const modals = document.querySelectorAll('.modal');
            modals.forEach(modal => {
                if (event.target == modal) {
                    modal.style.display = 'none';
                }
            });
        }

        // Close on Escape key
        document.addEventListener('keydown', function(event) {
            if (event.key === 'Escape') {
                const modals = document.querySelectorAll('.modal');
                modals.forEach(modal => {
                    modal.style.display = 'none';
                });
            }
        });
        </script>

        </body>
        </html>
        '''

        # Render using components
        import streamlit.components.v1 as components
        # Calculate dynamic height based on number of goals (tight fit)
        dynamic_height = 150 + (len(goals) * 50)  # Tight calculation: header + rows
        components.html(complete_html, height=dynamic_height, scrolling=False)
        
        # Calculate summary metrics
        total_target = sum(g.get('monthly_target', 0) for g in goals)
        total_achievement = sum(g.get('monthly_achievement', 0) for g in goals)
        avg_progress = calculate_progress(total_achievement, total_target)
        
        # Display metrics
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            render_metric_card("Total Target", f"{total_target:.2f}")
        with col2:
            render_metric_card("Total Achievement", f"{total_achievement:.2f}")
        with col3:
            render_metric_card("Progress", f"{avg_progress:.1f}%")
        
        st.markdown("---")
        
        # Action tabs: Update, Edit, Delete
        if st.session_state.get('viewing_employee_year'):
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        elif user['role'] in ['HR', 'Manager']:
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        else:
            action_tab1 = st.tabs(["📝 Update Achievements"])[0]
            action_tab2, action_tab3 = None, None
        
        # ===== UPDATE ACHIEVEMENTS TAB =====
        with action_tab1:
            st.subheader("Update Weekly Achievements")
            
            selected_goal_title = st.selectbox(
                "Select Goal to Update", 
                [g['goal_title'] for g in goals],
                key="update_goal_select"
            )
            selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)
            
            st.info("💡 **Note:** Weekly remarks will be automatically saved to respective week goal sheets")
            
            col1, col2 = st.columns(2)
            with col1:
                w1 = st.number_input("Week 1 Achievement", value=float(selected_goal.get('week1_achievement', 0)), key="w1_update")
                w2 = st.number_input("Week 2 Achievement", value=float(selected_goal.get('week2_achievement', 0)), key="w2_update")
            with col2:
                w3 = st.number_input("Week 3 Achievement", value=float(selected_goal.get('week3_achievement', 0)), key="w3_update")
                w4 = st.number_input("Week 4 Achievement", value=float(selected_goal.get('week4_achievement', 0)), key="w4_update")
            
            total_monthly = w1 + w2 + w3 + w4
            
            st.markdown("---")
            st.markdown(f"**📊 Monthly Achievement (Auto-calculated):** `{total_monthly:.2f}`")
            st.progress(min(total_monthly / selected_goal.get('monthly_target', 1), 1.0))
            
            st.markdown("---")
            st.markdown("**Weekly Remarks**")
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                w1_remarks = st.text_area("Week 1 Remarks", value=selected_goal.get('week1_remarks', ''), key="w1_remarks_update", height=100)
                w2_remarks = st.text_area("Week 2 Remarks", value=selected_goal.get('week2_remarks', ''), key="w2_remarks_update", height=100)
            with col_r2:
                w3_remarks = st.text_area("Week 3 Remarks", value=selected_goal.get('week3_remarks', ''), key="w3_remarks_update", height=100)
                w4_remarks = st.text_area("Week 4 Remarks", value=selected_goal.get('week4_remarks', ''), key="w4_remarks_update", height=100)
            
            if st.button("💾 Save Achievements", use_container_width=True, key="save_achievements"):
                updates = {
                    'week1_achievement': w1,
                    'week2_achievement': w2,
                    'week3_achievement': w3,
                    'week4_achievement': w4,
                    'monthly_achievement': total_monthly,
                    'week1_remarks': w1_remarks,
                    'week2_remarks': w2_remarks,
                    'week3_remarks': w3_remarks,
                    'week4_remarks': w4_remarks
                }
                
                if db.update_goal(selected_goal['goal_id'], updates):
                    st.success("✅ Achievements and remarks saved to monthly goal sheet!")
                    st.info("💡 Remarks are now available in the respective week views")
                    st.rerun()
                else:
                    st.error("❌ Failed to save achievements")
        
        # ===== EDIT GOAL TAB =====
        if action_tab2:
            with action_tab2:
                st.subheader("Edit Goal Details")
                
                edit_goal_title = st.selectbox(
                    "Select Goal to Edit", 
                    [g['goal_title'] for g in goals],
                    key="edit_goal_select"
                )
                edit_goal = next(g for g in goals if g['goal_title'] == edit_goal_title)
                
                with st.form("edit_goal_form"):
                    st.markdown("**Basic Information**")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        new_vertical = st.text_input("Vertical*", value=edit_goal.get('vertical', ''))
                        new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                        new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                    
                    with col2:
                        from datetime import datetime as dt
                        start_date_str = edit_goal.get('start_date', str(date.today()))
                        end_date_str = edit_goal.get('end_date', str(date.today()))
                        
                        if isinstance(start_date_str, str):
                            start_date_val = dt.strptime(start_date_str, '%Y-%m-%d').date()
                        else:
                            start_date_val = start_date_str
                        
                        if isinstance(end_date_str, str):
                            end_date_val = dt.strptime(end_date_str, '%Y-%m-%d').date()
                        else:
                            end_date_val = end_date_str
                        
                        new_start_date = st.date_input("Start Date", value=start_date_val)
                        new_end_date = st.date_input("End Date", value=end_date_val)
                        new_status = st.selectbox(
                            "Status", 
                            ['Active', 'Completed', 'On Hold', 'Cancelled'],
                            index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active'))
                        )
                    
                    new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
                    
                    st.markdown("**Targets**")
                    col3, col4 = st.columns(2)
                    
                    with col3:
                        new_monthly_target = st.number_input(
                            "Monthly Target*", 
                            min_value=0.0, 
                            value=float(edit_goal.get('monthly_target', 0))
                        )
                    
                    st.markdown("**Weekly Targets**")
                    col5, col6, col7, col8 = st.columns(4)
                    
                    with col5:
                        new_w1_target = st.number_input(
                            "Week 1 Target", 
                            min_value=0.0, 
                            value=float(edit_goal.get('week1_target', 0)),
                            key="edit_w1_target"
                        )
                    with col6:
                        new_w2_target = st.number_input(
                            "Week 2 Target", 
                            min_value=0.0, 
                            value=float(edit_goal.get('week2_target', 0)),
                            key="edit_w2_target"
                        )
                    with col7:
                        new_w3_target = st.number_input(
                            "Week 3 Target", 
                            min_value=0.0, 
                            value=float(edit_goal.get('week3_target', 0)),
                            key="edit_w3_target"
                        )
                    with col8:
                        new_w4_target = st.number_input(
                            "Week 4 Target", 
                            min_value=0.0, 
                            value=float(edit_goal.get('week4_target', 0)),
                            key="edit_w4_target"
                        )
                    
                    st.markdown("**Weekly Remarks**")
                    col_r5, col_r6, col_r7, col_r8 = st.columns(4)

                    with col_r5:
                        new_w1_remarks = st.text_area(
                            "Week 1 Remarks",
                            value=edit_goal.get('week1_remarks', ''),
                            key="edit_w1_remarks",
                            height=80
                        )
                    with col_r6:
                        new_w2_remarks = st.text_area(
                            "Week 2 Remarks",
                            value=edit_goal.get('week2_remarks', ''),
                            key="edit_w2_remarks",
                            height=80
                        )
                    with col_r7:
                        new_w3_remarks = st.text_area(
                            "Week 3 Remarks",
                            value=edit_goal.get('week3_remarks', ''),
                            key="edit_w3_remarks",
                            height=80
                        )
                    with col_r8:
                        new_w4_remarks = st.text_area(
                            "Week 4 Remarks",
                            value=edit_goal.get('week4_remarks', ''),
                            key="edit_w4_remarks",
                            height=80
                        )
                    
                    submitted = st.form_submit_button("💾 Save Changes", use_container_width=True)
                    
                    if submitted:
                        if new_vertical and new_title and new_kpi:
                            updates = {
                                'vertical': new_vertical,
                                'goal_title': new_title,
                                'goal_description': new_description,
                                'kpi': new_kpi,
                                'monthly_target': new_monthly_target,
                                'week1_target': new_w1_target,
                                'week2_target': new_w2_target,
                                'week3_target': new_w3_target,
                                'week4_target': new_w4_target,
                                'week1_remarks': new_w1_remarks,
                                'week2_remarks': new_w2_remarks,
                                'week3_remarks': new_w3_remarks,
                                'week4_remarks': new_w4_remarks,
                                'start_date': str(new_start_date),
                                'end_date': str(new_end_date),
                                'status': new_status
                            }
                            
                            if db.update_goal(edit_goal['goal_id'], updates):
                                st.success("✅ Goal updated successfully!")
                                st.rerun()
                            else:
                                st.error("❌ Failed to update goal")
                        else:
                            st.error("❌ Please fill all required fields (Vertical, Title, KPI)")
        
        # ===== DELETE GOAL TAB =====
        if action_tab3:
            with action_tab3:
                st.subheader("⚠️ Delete Goal")
                st.warning("**Warning:** Deleting a goal will also delete all associated feedback. This action cannot be undone!")
                
                delete_goal_title = st.selectbox(
                    "Select Goal to Delete", 
                    [g['goal_title'] for g in goals],
                    key="delete_goal_select"
                )
                delete_goal = next(g for g in goals if g['goal_title'] == delete_goal_title)
                
                st.markdown("**Goal Details:**")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.info(f"**Vertical:** {delete_goal.get('vertical', 'N/A')}")
                with col2:
                    st.info(f"**KPI:** {delete_goal.get('kpi', 'N/A')}")
                with col3:
                    st.info(f"**Target:** {delete_goal.get('monthly_target', 0)}")
                
                st.markdown(f"**Description:** {delete_goal.get('goal_description', 'No description')}")
                
                st.markdown("---")
                confirm_delete = st.checkbox("I understand this action cannot be undone", key="confirm_delete")
                
                col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
                
                with col_del2:
                    if st.button(
                        "🗑️ Delete Goal", 
                        disabled=not confirm_delete,
                        use_container_width=True,
                        type="primary"
                    ):
                        if db.delete_goal(delete_goal['goal_id']):
                            st.success("✅ Goal deleted successfully!")
                            st.balloons()
                            st.rerun()
                        else:
                            st.error("❌ Failed to delete goal")
    else:
        st.info("ℹ️ No goals yet for this month. Add your first goal below!")
        
        # Add new goal (only for own goals)
        if not st.session_state.get('viewing_employee_year'):
            st.markdown("---")
            display_add_goal_form(user, year, quarter, month)
    
    # Feedback section
    if goals:
        st.markdown("---")
        display_feedback_section(goals, 'month')

    
def display_assign_goal_form_monthly(user, year, quarter, month):
    """Form to assign goals in monthly view"""
    with st.form("assign_goal_monthly"):
        st.markdown(f"**Assigning for:** {get_month_name(month)} {year}, Q{quarter}")
        
        # Get employees based on role
        if user['role'] == 'HR':
            employees = [u for u in db.get_all_users() if u['role'] in ['Employee', 'Manager']]
        else:  # Manager
            employees = db.get_team_members(user['id'])
        
        if not employees:
            st.info("No employees available")
            return
        
        selected_emp = st.selectbox(
            "Assign To*",
            [f"{e['name']} ({e['email']})" for e in employees]
        )
        
        emp_email = selected_emp.split('(')[1].strip(')')
        emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
        
        col1, col2 = st.columns(2)
        with col1:
            vertical = st.text_input("Vertical*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        with col2:
            description = st.text_area("Description")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        st.markdown("**Weekly Targets**")
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, key="monthly_assign_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, key="monthly_assign_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, key="monthly_assign_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, key="monthly_assign_w4")
        
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Remarks (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_remarks = st.text_area("Week 1 Remarks", key="monthly_assign_w1_remarks", height=80)
        with col_r2:
            w2_remarks = st.text_area("Week 2 Remarks", key="monthly_assign_w2_remarks", height=80)
        with col_r3:
            w3_remarks = st.text_area("Week 3 Remarks", key="monthly_assign_w3_remarks", height=80)
        with col_r4:
            w4_remarks = st.text_area("Week 4 Remarks", key="monthly_assign_w4_remarks", height=80)
        # ========== END OF ADDED SECTION ==========

        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if vertical and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'vertical': vertical,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_remarks,  # ✅ ADD
                    'week2_remarks': w2_remarks,  # ✅ ADD
                    'week3_remarks': w3_remarks,  # ✅ ADD
                    'week4_remarks': w4_remarks,
                    'start_date': f'{year}-{month:02d}-01',
                    'end_date': f'{year}-{month:02d}-28',
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    st.success(f"✅ Goal assigned to {selected_emp}")
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")


# ============================================
# ENHANCED FEEDBACK HISTORY (WITH USER INFO)
# ============================================
def display_feedback_history():
    """Display feedback history with reply and new feedback options"""
    user = st.session_state.user

    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    role = user['role']
    
    st.title("💬 Feedback History")
    
    # Add New Feedback Button
    if st.button("➕ Add New Feedback", use_container_width=True):
        st.session_state.adding_new_feedback = True
    
    # New Feedback Modal
    if st.session_state.get('adding_new_feedback'):
        st.markdown("---")
        st.subheader("➕ Add New Feedback")
        
        with st.form("add_new_feedback_form"):
            # Get all goals based on role
            if role == 'HR':
                all_users = db.get_all_users()
                all_goals = []
                for u in all_users:
                    user_goals = db.get_user_all_goals(u['id'])
                    for g in user_goals:
                        g['user_name'] = u['name']
                    all_goals.extend(user_goals)
            elif role == 'Manager':
                team_members = db.get_team_members(user['id'])
                all_goals = []
                for m in team_members:
                    member_goals = db.get_user_all_goals(m['id'])
                    for g in member_goals:
                        g['user_name'] = m['name']
                    all_goals.extend(member_goals)
                # Add own goals
                own_goals = db.get_user_all_goals(user['id'])
                for g in own_goals:
                    g['user_name'] = user['name']
                all_goals.extend(own_goals)
            else:
                all_goals = db.get_user_all_goals(user['id'])
                for g in all_goals:
                    g['user_name'] = user['name']
            
            if all_goals:
                selected_goal_str = st.selectbox(
                    "Select Goal*",
                    [f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals]
                )
                
                selected_goal = all_goals[[f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals].index(selected_goal_str)]
                
                # Determine feedback type
                if role == 'HR':
                    fb_types = ["HR", "Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "HR"]
                elif role == 'Manager':
                    fb_types = ["Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "Manager"]
                else:
                    fb_types = ["Self Appraisal"]
                
                fb_type = st.selectbox("Feedback Type*", fb_types)
                rating = st.slider("Rating", 1, 5, 3)
                comment = st.text_area("Comment*")
                
                col_submit, col_cancel = st.columns(2)
                
                with col_submit:
                    if st.form_submit_button("Submit Feedback", use_container_width=True):
                        if comment.strip():
                            feedback_data = {
                                'goal_id': selected_goal['goal_id'],
                                'user_id': selected_goal['user_id'],
                                'feedback_by': user['id'],
                                'feedback_type': fb_type,
                                'rating': rating,
                                'comment': comment.strip(),
                                'level': 'general'
                            }
                            if db.create_feedback(feedback_data):
                                st.success("✅ Feedback submitted!")
                                del st.session_state.adding_new_feedback
                                st.rerun()
                        else:
                            st.error("❌ Please enter a comment")
                
                with col_cancel:
                    if st.form_submit_button("Cancel", use_container_width=True):
                        del st.session_state.adding_new_feedback
                        st.rerun()
            else:
                st.info("No goals available for feedback")
                if st.form_submit_button("Cancel", use_container_width=True):
                    del st.session_state.adding_new_feedback
                    st.rerun()
        
        st.markdown("---")
    
    # Get feedback based on role
    if role == 'HR':
        all_feedbacks = db.get_all_feedback()
        st.subheader("All Feedback (System-wide)")
    elif role == 'Manager':
        my_feedback = db.get_user_all_feedback(user['id'])
        team_feedback = []
        for member in db.get_team_members(user['id']):
            team_feedback.extend(db.get_user_all_feedback(member['id']))
        all_feedbacks = my_feedback + team_feedback
        st.subheader("My Feedback & Team Feedback")
    else:
        all_feedbacks = db.get_user_all_feedback(user['id'])
        st.subheader("My Feedback")
    
    all_feedbacks.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    if not all_feedbacks:
        st.info("No feedback history found")
        return
    
    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_type = st.selectbox("Filter by Type", ["All", "Self Appraisal", "Manager", "HR"])
    with col2:
        filter_rating = st.selectbox("Filter by Rating", ["All", "5", "4", "3", "2", "1"])
    with col3:
        search_goal = st.text_input("Search Goal")
    
    # Filter feedbacks
    filtered = all_feedbacks
    if filter_type != "All":
        filtered = [f for f in filtered if f.get('feedback_type') == filter_type]
    if filter_rating != "All":
        filtered = [f for f in filtered if f.get('rating') == int(filter_rating)]
    if search_goal:
        filtered = [f for f in filtered if search_goal.lower() in f.get('goal_title', '').lower()]
    
    # Display feedback
    st.markdown(f"**Showing {len(filtered)} of {len(all_feedbacks)} feedback entries**")
    
    for feedback in filtered:
        # Get user who received feedback
        feedback_user = db.get_user_by_id(feedback.get('user_id'))
        feedback_user_name = feedback_user['name'] if feedback_user else 'Unknown'
        
        is_new = False
        created_at_str = feedback.get('created_at', '')
        if created_at_str:
            try:
                utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                utc_time = pytz.utc.localize(utc_time)
                ist_time = utc_time.astimezone(IST)
                hours_ago = (datetime.now(IST) - ist_time).total_seconds() / 3600
                is_new = hours_ago <= 24
            except:
                pass

        with st.container():
            # Header with user info
            st.markdown(f"### 💬 Feedback for: **{feedback_user_name}**")
            
            col_fb1, col_fb2 = st.columns([3, 1])
            
            with col_fb1:
                st.markdown(f"**Goal:** {feedback.get('goal_title', 'N/A')}")
                st.markdown(f"**Type:** {feedback.get('feedback_type')} | **Rating:** {'⭐' * feedback.get('rating', 0)}")
                st.markdown(f"**Comment:** {feedback.get('comment')}")
                
                # Reply Section
                replies = db.get_feedback_replies(feedback.get('feedback_id'))
                if replies:
                    st.markdown("**Replies:**")
                    for reply in replies:
                        st.info(f"↳ **{reply.get('reply_by_name', 'Unknown')}:** {reply.get('reply_text')} ({reply.get('created_at', '')[:16]})")
                
                # Add Reply Button
                if st.button("💬 Reply", key=f"reply_btn_{feedback.get('feedback_id')}"):
                    st.session_state[f"replying_to_{feedback.get('feedback_id')}"] = True
                
                # Reply Form
                if st.session_state.get(f"replying_to_{feedback.get('feedback_id')}"):
                    with st.form(f"reply_form_{feedback.get('feedback_id')}"):
                        reply_text = st.text_area("Your Reply*", key=f"reply_text_{feedback.get('feedback_id')}")
                        
                        col_reply1, col_reply2 = st.columns(2)
                        with col_reply1:
                            if st.form_submit_button("Send Reply"):
                                if reply_text.strip():
                                    reply_data = {
                                        'feedback_id': feedback.get('feedback_id'),
                                        'reply_by': user['id'],
                                        'reply_text': reply_text.strip()
                                    }
                                    if db.create_feedback_reply(reply_data):
                                        st.success("✅ Reply added!")
                                        del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                        st.rerun()
                                else:
                                    st.error("❌ Please enter a reply")
                        
                        with col_reply2:
                            if st.form_submit_button("Cancel"):
                                del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                st.rerun()
            
            with col_fb2:
                st.markdown(f"**Given By:** {feedback.get('feedback_by_name', 'Unknown')}")
                
                # Convert UTC to IST
                created_at_str = feedback.get('created_at', '')
                if created_at_str:
                    try:
                        utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                        utc_time = pytz.utc.localize(utc_time)
                        ist_time = utc_time.astimezone(IST)
                        st.markdown(f"**Date:** {ist_time.strftime('%Y-%m-%d')}")
                        st.markdown(f"**Time:** {ist_time.strftime('%I:%M %p IST')}")
                    except:
                        st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                else:
                    st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                
                # Delete button for HR
                if role == 'HR':
                    if st.button("🗑️ Delete", key=f"del_fb_{feedback.get('feedback_id')}", use_container_width=True):
                        if db.delete_feedback(feedback.get('feedback_id')):
                            st.success("✅ Feedback deleted!")
                            st.rerun()
            
            st.markdown("---")




# ============================================
# WEEK VIEW (Keep existing)
# ============================================
def display_week_view(user, year, quarter, month, week_num):
    """Display week-specific view with remarks from monthly goals"""
    if not year or not quarter or not month or not week_num:
        st.warning("⚠️ Invalid week view parameters. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()
    month_name = get_month_name(month)
    st.subheader(f" Week {week_num} - {month_name} {year}")
    
    # Get monthly goals to show breakdown
    monthly_goals = db.get_month_goals(user['id'], year, quarter, month)
    
    if monthly_goals:
        st.markdown("**Weekly Breakdown from Monthly Goals**")
        
        table_data = []
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement = goal.get(f'week{week_num}_achievement', 0)
            week_remarks = goal.get(f'week{week_num}_remarks', '')
            progress = calculate_progress(week_achievement, week_target)
            
            table_data.append({
                'Vertical': goal.get('vertical', ''),
                'Title': goal['goal_title'],
                'KPI': goal.get('kpi', ''),
                'Target': week_target,
                'Achievement': week_achievement,
                'Progress %': f"{progress:.1f}%",
                'Remarks': week_remarks if week_remarks else '-'
            })
        
        df = pd.DataFrame(table_data)
        st.dataframe(df, use_container_width=True)
        
        # Show detailed goal cards with remarks
        st.markdown("---")
        st.markdown("**Goal Details**")
        
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement = goal.get(f'week{week_num}_achievement', 0)
            week_remarks = goal.get(f'week{week_num}_remarks', '')
            progress = calculate_progress(week_achievement, week_target)
            
            with st.expander(f"📌 {goal['goal_title']}"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Target", week_target)
                with col2:
                    st.metric("Achievement", week_achievement)
                with col3:
                    st.metric("Progress", f"{progress:.1f}%")
                
                render_progress_bar(progress, goal['goal_title'])
                
                st.markdown(f"**Vertical:** {goal.get('vertical', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                
                # Show remarks
                st.markdown("---")
                if week_remarks:
                    st.markdown(f"**📝 Week {week_num} Remarks:**")
                    st.info(week_remarks)
                else:
                    st.caption(f"No remarks added for Week {week_num} yet")
    
    # Week-specific goals with management
    st.markdown("---")
    week_goals = db.get_week_goals(user['id'], year, quarter, month, week_num)
    
    if week_goals:
        st.markdown("**Week-Specific Goals**")
        
        # Tabs for managing week goals
        tab1, tab2, tab3 = st.tabs(["📋 View Goals", "✏️ Edit Goal", "🗑️ Delete Goal"])
        
        with tab1:
            for goal in week_goals:
                with st.expander(f"📌 {goal['goal_title']}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Target", goal.get('weekly_target', 0))
                    with col2:
                        st.metric("Achievement", goal.get('weekly_achievement', 0))
                    with col3:
                        progress = calculate_progress(
                            goal.get('weekly_achievement', 0),
                            goal.get('weekly_target', 0)
                        )
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown(f"**Vertical:** {goal.get('vertical', 'N/A')}")
                    st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                    st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                    
                    # Show remarks for week-specific goals
                    week_remarks = goal.get(f'week{week_num}_remarks', '')
                    if week_remarks:
                        st.markdown(f"**📝 Remarks:** {week_remarks}")
                    else:
                        st.caption("No remarks added yet")

        with tab2:
            st.subheader("Edit Week-Specific Goal")
            
            edit_week_goal_title = st.selectbox(
                "Select Goal to Edit",
                [g['goal_title'] for g in week_goals],
                key=f"edit_week_goal_{week_num}"
            )
            edit_week_goal = next(g for g in week_goals if g['goal_title'] == edit_week_goal_title)
            
            with st.form(f"edit_week_goal_form_{week_num}"):
                col1, col2 = st.columns(2)
                with col1:
                    new_vertical = st.text_input("Vertical", value=edit_week_goal.get('vertical', ''))
                    new_title = st.text_input("Goal Title*", value=edit_week_goal['goal_title'])
                with col2:
                    new_kpi = st.text_input("KPI", value=edit_week_goal.get('kpi', ''))
                    new_status = st.selectbox(
                        "Status",
                        ['Active', 'Completed', 'On Hold', 'Cancelled'],
                        index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_week_goal.get('status', 'Active'))
                    )
                
                new_description = st.text_area("Description", value=edit_week_goal.get('goal_description', ''))
                
                col3, col4 = st.columns(2)
                with col3:
                    new_target = st.number_input("Weekly Target", min_value=0.0, value=float(edit_week_goal.get('weekly_target', 0)))
                with col4:
                    new_achievement = st.number_input("Weekly Achievement", min_value=0.0, value=float(edit_week_goal.get('weekly_achievement', 0)))
                
                # Add remarks field for week-specific goals
                new_remarks = st.text_area(f"Week {week_num} Remarks", value=edit_week_goal.get(f'week{week_num}_remarks', ''), height=100)
                
                if st.form_submit_button(" Save Changes", use_container_width=True):
                    if new_title:
                        updates = {
                            'vertical': new_vertical,
                            'goal_title': new_title,
                            'goal_description': new_description,
                            'kpi': new_kpi,
                            'weekly_target': new_target,
                            'weekly_achievement': new_achievement,
                            'status': new_status,
                            f'week{week_num}_remarks': new_remarks
                        }
                        
                        if db.update_goal(edit_week_goal['goal_id'], updates):
                            st.success("✅ Week goal updated!")
                            st.rerun()
                    else:
                        st.error("❌ Goal title is required")
        
        with tab3:
            st.subheader("⚠️ Delete Week-Specific Goal")
            st.warning("This action cannot be undone!")
            
            delete_week_goal_title = st.selectbox(
                "Select Goal to Delete",
                [g['goal_title'] for g in week_goals],
                key=f"delete_week_goal_{week_num}"
            )
            delete_week_goal = next(g for g in week_goals if g['goal_title'] == delete_week_goal_title)
            
            st.info(f"**Goal:** {delete_week_goal['goal_title']}")
            st.info(f"**Target:** {delete_week_goal.get('weekly_target', 0)}")
            
            confirm = st.checkbox("I understand this cannot be undone", key=f"confirm_week_del_{week_num}")
            
            if st.button("🗑️ Delete Goal", disabled=not confirm, use_container_width=True):
                if db.delete_goal(delete_week_goal['goal_id']):
                    st.success("✅ Week goal deleted!")
                    st.rerun()
    
    # Add week-specific goal (only for own goals)
    if not st.session_state.get('viewing_employee_year'):
        
        with st.expander("➕ Add Week-Specific Goal"):
            with st.form(f"add_week{week_num}_goal"):
                vertical = st.text_input("Vertical")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI")
                target = st.number_input("Weekly Target", min_value=0.0)
                description = st.text_area("Description")
                remarks = st.text_area(f"Week {week_num} Remarks", height=100)
                
                if st.form_submit_button("Create Goal"):
                    if title:
                        goal_data = {
                            'user_id': user['id'],
                            'year': year,
                            'quarter': quarter,
                            'month': month,
                            'week': week_num,
                            'vertical': vertical,
                            'goal_title': title,
                            'goal_description': description,
                            'kpi': kpi,
                            'weekly_target': target,
                            f'week{week_num}_remarks': remarks,
                            'start_date': str(date.today()),
                            'end_date': str(date.today())
                        }
                        if db.create_goal(goal_data):
                            st.success("✅ Goal created!")
                            st.rerun()
                    else:
                        st.error("❌ Please enter a goal title")
    
    # Week feedback
    st.markdown("---")
    all_week_goals = monthly_goals + week_goals
    if all_week_goals:
        display_feedback_section(all_week_goals, f'week{week_num}')

def display_add_goal_form(user, year, quarter, month):
    """Display form to add new goal"""
    with st.expander("➕ Add New Monthly Goal"):
        with st.form("add_goal"):
            col1, col2 = st.columns(2)
            
            with col1:
                vertical = st.text_input("Vertical*")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI*")
                monthly_target = st.number_input("Monthly Target*", min_value=0.0)
            
            with col2:
                start_date = st.date_input("Start Date", value=date.today())
                end_date = st.date_input("End Date", value=date.today())
                description = st.text_area("Description")
            
            st.markdown("**Weekly Targets**")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0)
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0)
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0)
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0)

            # ADD THIS NEW SECTION
            st.markdown("**Weekly Remarks (Optional)**")
            col_r1, col_r2, col_r3, col_r4 = st.columns(4)
            with col_r1:
                w1_remarks = st.text_area("Week 1 Remarks", key="w1_remarks_create", height=80)
            with col_r2:
                w2_remarks = st.text_area("Week 2 Remarks", key="w2_remarks_create", height=80)
            with col_r3:
                w3_remarks = st.text_area("Week 3 Remarks", key="w3_remarks_create", height=80)
            with col_r4:
                w4_remarks = st.text_area("Week 4 Remarks", key="w4_remarks_create", height=80)
                        
            if st.form_submit_button("Create Goal", use_container_width=True):
                if vertical and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'vertical': vertical,
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_remarks,  # ADD
                        'week2_remarks': w2_remarks,  # ADD
                        'week3_remarks': w3_remarks,  # ADD
                        'week4_remarks': w4_remarks,
                        'start_date': str(start_date),
                        'end_date': str(end_date)
                    }
                    
                    is_valid, error_msg = validate_goal_data(goal_data)
                    if is_valid:
                        if db.create_goal(goal_data):
                            st.success("✅ Goal created successfully!")
                            st.rerun()
                    else:
                        st.error(f"❌ {error_msg}")
                else:
                    st.error("❌ Please fill all required fields (Vertical, Title, KPI)")

def display_add_goal_form_inline(user, year, quarter, month):
    """Inline form to add new goal at the top"""
    with st.form("add_goal_inline_top"):
        col1, col2 = st.columns(2)
        
        with col1:
            vertical = st.text_input("Vertical*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        with col2:
            start_date = st.date_input("Start Date", value=date(year, month, 1))
            end_date = st.date_input("End Date", value=date(year, month, calendar.monthrange(year, month)[1]))
            description = st.text_area("Description", height=100)
        
        st.markdown("**Weekly Targets**")
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, key="inline_top_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, key="inline_top_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, key="inline_top_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, key="inline_top_w4")
        
        st.markdown("**Weekly Remarks (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_remarks = st.text_area("Week 1 Remarks", key="inline_top_w1_remarks", height=80)
        with col_r2:
            w2_remarks = st.text_area("Week 2 Remarks", key="inline_top_w2_remarks", height=80)
        with col_r3:
            w3_remarks = st.text_area("Week 3 Remarks", key="inline_top_w3_remarks", height=80)
        with col_r4:
            w4_remarks = st.text_area("Week 4 Remarks", key="inline_top_w4_remarks", height=80)
        
        col_submit, col_cancel = st.columns(2)
        
        with col_submit:
            if st.form_submit_button("✅ Create Goal", use_container_width=True):
                if vertical and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'vertical': vertical,
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_remarks,
                        'week2_remarks': w2_remarks,
                        'week3_remarks': w3_remarks,
                        'week4_remarks': w4_remarks,
                        'start_date': str(start_date),
                        'end_date': str(end_date),
                        'created_by': user['id']  # Track who created the goal
                    }
                    
                    if db.create_goal(goal_data):
                        st.success("✅ Goal created successfully!")
                        st.session_state.show_create_goal_form = False
                        st.rerun()
                    else:
                        st.error("❌ Failed to create goal")
                else:
                    st.error("❌ Please fill all required fields (Vertical, Title, KPI)")
        
        with col_cancel:
            if st.form_submit_button("❌ Cancel", use_container_width=True):
                st.session_state.show_create_goal_form = False
                st.rerun()

def display_feedback_section(goals, level):
    """Display feedback section"""
    user = st.session_state.user
    
    st.subheader("💬 Feedback & Appraisals")
    
    if not goals:
        st.info("ℹ️ No goals to provide feedback on")
        return
    
    selected_goal_title = st.selectbox("Select Goal for Feedback", 
                                       [g['goal_title'] for g in goals], 
                                       key=f"fb_{level}")
    selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)
    
    # Get feedback
    feedbacks = db.get_goal_feedback(selected_goal['goal_id'])
    
    # Separate by type
    manager_fb = [f for f in feedbacks if f.get('feedback_type') == 'Manager']
    hr_fb = [f for f in feedbacks if f.get('feedback_type') == 'HR']
    self_fb = [f for f in feedbacks if f.get('feedback_type') == 'Self Appraisal']
    
    # Display in columns
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**Manager Feedback**")
        if manager_fb:
            for fb in manager_fb:
                render_feedback_card(fb, 'Manager')
        else:
            st.caption("No manager feedback yet")
    
    with col2:
        st.markdown("** HR Feedback**")
        if hr_fb:
            for fb in hr_fb:
                render_feedback_card(fb, 'HR')
        else:
            st.caption("No HR feedback yet")
    
    with col3:
        st.markdown("** Self Appraisal**")
        if self_fb:
            for fb in self_fb:
                render_feedback_card(fb, 'Self Appraisal')
        else:
            st.caption("No self appraisal yet")
    
    # Add feedback form (only if not viewing employee goals or if manager/HR viewing their team)
    can_add_feedback = True
    if st.session_state.get('viewing_employee_year'):
        # Manager or HR can add feedback to team member goals
        if user['role'] in ['Manager', 'HR']:
            can_add_feedback = True
        else:
            can_add_feedback = False
    
    if can_add_feedback:
        st.markdown("---")
        with st.expander("➕ Add Feedback"):
            with st.form(f"feedback_form_{selected_goal['goal_id']}_{level}"):
                # Determine feedback types based on role
                if user['role'] == 'HR':
                    fb_types = ["HR", "Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "HR"]
                elif user['role'] == 'Manager':
                    fb_types = ["Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "Manager"]
                else:
                    fb_types = ["Self Appraisal"]
                
                fb_type = st.selectbox("Feedback Type", fb_types)
                rating = st.slider("Rating", 1, 5, 3)
                comment = st.text_area("Comment*")
                
                if st.form_submit_button("Submit Feedback", use_container_width=True):
                    if comment.strip():
                        feedback_data = {
                            'goal_id': selected_goal['goal_id'],
                            'user_id': selected_goal['user_id'],
                            'feedback_by': user['id'],
                            'feedback_type': fb_type,
                            'rating': rating,
                            'comment': comment.strip(),
                            'level': level
                        }
                        if db.create_feedback(feedback_data):
                            st.success("✅ Feedback submitted!")
                            st.rerun()
                    else:
                        st.error("❌ Please enter a comment")


# Keep display_profile, display_permissions, display_employee_management from previous code

def display_profile():
    """Display and edit user profile"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    st.title("👤 My Profile")
    
    tab1, tab2 = st.tabs(["📝 Edit Profile", "🔒 Change Password"])
    
    with tab1:
        with st.form("edit_profile"):
            st.subheader("Personal Information")
            
            col1, col2 = st.columns(2)
            with col1:
                new_name = st.text_input("Full Name*", value=user['name'])
                new_email = st.text_input("Email*", value=user['email'])
            
            with col2:
                new_designation = st.text_input("Designation", value=user.get('designation', ''))
                new_department = st.text_input("Department", value=user.get('department', ''))
            
            if st.form_submit_button("💾 Save Changes", use_container_width=True):
                if new_name and new_email:
                    updates = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'designation': new_designation,
                        'department': new_department
                    }
                    
                    if db.update_user(user['id'], updates):
                        st.session_state.user.update(updates)
                        st.success("✅ Profile updated successfully!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to update profile")
                else:
                    st.error("❌ Name and Email are required")
    
    with tab2:
        # Initialize session state for forgot password in profile
        if 'profile_forgot_password' not in st.session_state:
            st.session_state.profile_forgot_password = False
        if 'profile_reset_token_sent' not in st.session_state:
            st.session_state.profile_reset_token_sent = False
        
        # Track password input for strength meter (OUTSIDE FORM)
        if 'temp_new_password' not in st.session_state:
            st.session_state.temp_new_password = ""
        
        # ===== NORMAL CHANGE PASSWORD =====
        if not st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("Change Password")
            
            old_password = st.text_input("Current Password*", type="password", key="old_pass_input")
            new_password = st.text_input("New Password*", type="password", key="new_pass_input")
            
            # Update temp password for strength meter
            if new_password != st.session_state.temp_new_password:
                st.session_state.temp_new_password = new_password
            
            # Password strength meter (OUTSIDE FORM)
            if new_password:
                render_password_strength_meter(new_password, "profile_change")
            
            confirm_password = st.text_input("Confirm New Password*", type="password", key="confirm_pass_input")
            
            col_submit, col_forgot = st.columns(2)
            
            with col_submit:
                if st.button("🔒 Change Password", use_container_width=True, key="change_pass_btn"):
                    if old_password == user['password']:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if db.update_user(user['id'], {'password': new_password}):
                                    st.session_state.user['password'] = new_password
                                    st.session_state.temp_new_password = ""
                                    st.success("✅ Password changed successfully!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to change password")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.error("❌ Current password is incorrect")
            
            with col_forgot:
                if st.button("🔑 Forgot Password?", use_container_width=True, key="forgot_pass_btn"):
                    st.session_state.profile_forgot_password = True
                    st.rerun()
        
        # ===== FORGOT PASSWORD - REQUEST TOKEN =====
        elif st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("🔑 Reset Password via Email")
            st.info("We'll send a reset token to your registered email address.")
            
            st.markdown(f"**Your Email:** {user['email']}")
            st.caption("A reset token will be sent to this email address")
            
            col_send, col_cancel = st.columns(2)
            
            with col_send:
                if st.button("📧 Send Reset Token", use_container_width=True, key="send_token_btn"):
                    # Generate token
                    token = db.create_password_reset_token(user['email'])
                    
                    if token:
                        # Try to send email
                        email_sent = send_password_reset_email(user['email'], token)
                        
                        if email_sent:
                            st.success("✅ Reset token sent to your email!")
                            st.info(f"**Backup Token:** `{token}`\n\n(In case you don't receive the email)")
                        else:
                            st.warning("⚠️ Could not send email. Please use this token:")
                            st.code(token, language=None)
                        
                        st.session_state.profile_reset_token_sent = True
                        st.rerun()
                    else:
                        st.error("❌ Failed to generate reset token")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_forgot_btn"):
                    st.session_state.profile_forgot_password = False
                    st.rerun()
        
        # ===== RESET PASSWORD WITH TOKEN =====
        elif st.session_state.profile_reset_token_sent:
            st.subheader("🔒 Enter Reset Token")
            st.info("Check your email for the reset token and enter it below with your new password.")
            
            reset_token = st.text_input("🎫 Reset Token*", placeholder="Enter 8-character token", key="reset_token_input")
            new_password_reset = st.text_input("🔒 New Password*", type="password", placeholder="••••••••", key="new_pass_reset_input")
            
            # Password strength meter (OUTSIDE FORM)
            if new_password_reset:
                render_password_strength_meter(new_password_reset, "profile_reset")
            
            confirm_password_reset = st.text_input("🔒 Confirm New Password*", type="password", placeholder="••••••••", key="confirm_pass_reset_input")
            
            col_reset, col_cancel = st.columns(2)
            
            with col_reset:
                if st.button("✅ Reset Password", use_container_width=True, key="reset_pass_btn"):
                    if reset_token and new_password_reset and confirm_password_reset:
                        if new_password_reset == confirm_password_reset:
                            if len(new_password_reset) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password_reset)
                                
                                # Reset password using token
                                if db.reset_password_with_token(reset_token, new_password_reset):
                                    st.success("✅ Password reset successful!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.balloons()
                                    
                                    # Update session
                                    st.session_state.user['password'] = new_password_reset
                                    st.session_state.profile_forgot_password = False
                                    st.session_state.profile_reset_token_sent = False
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token. Please request a new one.")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_reset_btn"):
                    st.session_state.profile_forgot_password = False
                    st.session_state.profile_reset_token_sent = False
                    st.rerun()
            
            # Option to resend token
            st.markdown("---")
            if st.button("📧 Resend Reset Token", use_container_width=True, key="resend_token_btn"):
                token = db.create_password_reset_token(user['email'])
                if token:
                    email_sent = send_password_reset_email(user['email'], token)
                    if email_sent:
                        st.success("✅ New token sent to your email!")
                    st.info(f"**Backup Token:** `{token}`")
                else:
                    st.error("❌ Failed to generate new token")
                    
def display_permissions():
    """Manage user permissions (HR only)"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    if user['role'] != 'HR':
        st.warning("⚠️ Only HR can access this page")
        return
    
    st.title("🔐 Permissions Management")
    
    st.info("**Note:** This section allows HR to manage granular permissions for users")
    
    all_users = db.get_all_users()
    
    # Permission categories
    permissions = {
        'view_all_goals': 'View All Goals',
        'edit_all_goals': 'Edit All Goals',
        'delete_all_goals': 'Delete All Goals',
        'view_all_feedback': 'View All Feedback',
        'edit_all_feedback': 'Edit All Feedback',
        'delete_all_feedback': 'Delete All Feedback',
        'create_users': 'Create Users',
        'edit_users': 'Edit Users',
        'delete_users': 'Delete Users',
        'manage_teams': 'Manage Teams',
        'view_analytics': 'View Analytics',
        'export_data': 'Export Data'
    }
    
    selected_user = st.selectbox(
        "Select User",
        [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users if u['id'] != user['id']]
    )
    
    if selected_user:
        user_email = selected_user.split(' - ')[1]
        selected_user_obj = next(u for u in all_users if u['email'] == user_email)
        
        st.subheader(f"Permissions for {selected_user_obj['name']}")
        
        # Get current permissions
        current_perms = db.get_user_permissions(selected_user_obj['id'])
        
        with st.form("update_permissions"):
            st.markdown("**Grant Permissions:**")
            
            selected_perms = []
            cols = st.columns(2)
            for idx, (perm_key, perm_label) in enumerate(permissions.items()):
                with cols[idx % 2]:
                    if st.checkbox(perm_label, value=perm_key in current_perms, key=perm_key):
                        selected_perms.append(perm_key)
            
            if st.form_submit_button("💾 Update Permissions", use_container_width=True):
                if db.update_user_permissions(selected_user_obj['id'], selected_perms):
                    st.success(f"✅ Permissions updated for {selected_user_obj['name']}")
                    st.rerun()
                else:
                    st.error("❌ Failed to update permissions")
        
        # Show current permissions
        st.markdown("---")
        st.subheader("Current Permissions")
        if current_perms:
            for perm in current_perms:
                st.success(f"✓ {permissions.get(perm, perm)}")
        else:
            st.info("No special permissions granted")


def display_employee_management():
    """Employee management - create users, teams"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    if user['role'] != 'HR':
        st.warning("⚠️ Only HR can access this page")
        return
    
    st.title("⚙️ Employee Management")
    
    tab1, tab2, tab3 = st.tabs(["👤 Create Employee", "👥 Manage Teams", "📋 View All Teams"])
    
    # ===== CREATE EMPLOYEE TAB =====
    with tab1:
        st.subheader("Create New Employee")
        
        with st.form("create_employee"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*")
                new_email = st.text_input("Email*")
                new_password = st.text_input("Password*", type="password")
                new_designation = st.text_input("Designation")
            
            with col2:
                new_role = st.selectbox("Role*", ["Employee", "Manager", "HR"])
                new_department = st.text_input("Department")
                
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                selected_manager = st.selectbox("Assign to Manager", manager_options)
            
            if st.form_submit_button("Create Employee", use_container_width=True):
                if new_name and new_email and new_password and new_role:
                    manager_id = None
                    if selected_manager != "None":
                        manager_email = selected_manager.split('(')[1].strip(')')
                        manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                    
                    employee_data = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'password': new_password,
                        'designation': new_designation,
                        'role': new_role,
                        'department': new_department,
                        'manager_id': manager_id
                    }
                    
                    if db.create_user(employee_data):
                        st.success(f"✅ Employee {new_name} created successfully!")
                        st.balloons()
                    else:
                        st.error("❌ Failed to create employee")
                else:
                    st.error("❌ Please fill all required fields")
    
    # ===== MANAGE TEAMS TAB =====
    with tab2:
        st.subheader("Assign Employees to Managers")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        employees = [u for u in db.get_all_users() if u['role'] == 'Employee']
        
        if not managers:
            st.info("No managers found. Create a manager first.")
        elif not employees:
            st.info("No employees found. Create employees first.")
        else:
            selected_manager_name = st.selectbox(
                "Select Manager",
                [f"{m['name']} ({m['email']})" for m in managers]
            )
            
            manager_email = selected_manager_name.split('(')[1].strip(')')
            selected_manager = next(m for m in managers if m['email'] == manager_email)
            
            current_team = db.get_team_members(selected_manager['id'])
            st.write(f"**Current Team Size:** {len(current_team)}")
            
            unassigned_employees = [e for e in employees if not e.get('manager_id')]
            
            if unassigned_employees:
                selected_employees = st.multiselect(
                    "Select Employees to Add",
                    [f"{e['name']} ({e['email']})" for e in unassigned_employees]
                )
                
                if st.button("Assign to Team", use_container_width=True):
                    for emp_str in selected_employees:
                        emp_email = emp_str.split('(')[1].strip(')')
                        emp_id = next(e['id'] for e in unassigned_employees if e['email'] == emp_email)
                        db.update_user(emp_id, {'manager_id': selected_manager['id']})
                    
                    st.success(f"✅ Assigned {len(selected_employees)} employees to {selected_manager['name']}")
                    st.rerun()
            else:
                st.info("No unassigned employees available")
    
    # ===== VIEW ALL TEAMS TAB =====
    with tab3:
        st.subheader("All Teams Overview")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        for manager in managers:
            with st.expander(f"👔 {manager['name']}'s Team ({manager.get('department', 'N/A')})"):
                team = db.get_team_members(manager['id'])
                
                if team:
                    team_data = []
                    for member in team:
                        team_data.append({
                            'Name': member['name'],
                            'Email': member['email'],
                            'Designation': member.get('designation', 'N/A'),
                            'Department': member.get('department', 'N/A')
                        })
                    
                    df = pd.DataFrame(team_data)
                    st.dataframe(df, use_container_width=True)
                    
                    # Option to remove members
                    remove_member = st.selectbox(
                        "Remove Member",
                        ["None"] + [m['name'] for m in team],
                        key=f"remove_{manager['id']}"
                    )
                    
                    if remove_member != "None" and st.button(f"Remove {remove_member}", key=f"btn_remove_{manager['id']}"):
                        member_id = next(m['id'] for m in team if m['name'] == remove_member)
                        db.update_user(member_id, {'manager_id': None})
                        st.success(f"✅ Removed {remove_member} from team")
                        st.rerun()
                else:
                    st.info("No team members yet")


# ============================================
# SIDEBAR
# ============================================
def render_sidebar():
    """Render sidebar with navigation"""
    user = st.session_state.user
    role = user['role']
    
    with st.sidebar:
        # User profile
        render_user_avatar(user)
        
        st.markdown("---")
        
        # Navigation Menu
        st.markdown(" Menu")
        
        if st.button(" Dashboard", use_container_width=True, key="nav_dashboard"):
            st.session_state.page = 'dashboard'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        if st.button(" My Goals", use_container_width=True, key="nav_my_goals"):
            st.session_state.page = 'my_goals'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        if st.button(" View All Goals", use_container_width=True, key="nav_view_all_goals"):
            st.session_state.page = 'view_all_goals'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        

        # Role-specific navigation
        if role in ['HR', 'Manager']:
            if st.button(" Employees", use_container_width=True, key="nav_employees"):
                st.session_state.page = 'employees'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        if role == 'HR':
            if st.button(" Employee Management", use_container_width=True, key="nav_emp_mgmt"):
                st.session_state.page = 'employee_management'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
            
            if st.button("HR Info", use_container_width=True, key="nav_hr_info"):
                st.session_state.page = 'hr_info'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        
        if st.button(" Analytics", use_container_width=True, key="nav_analytics"):
            st.session_state.page = 'analytics'
            save_session_to_storage()
            st.rerun()

        if st.button(" Feedback History", use_container_width=True, key="nav_feedback"):
            st.session_state.page = 'feedback_history'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        if role == 'HR':
            if st.button(" Permissions", use_container_width=True, key="nav_permissions"):
                st.session_state.page = 'permissions'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        # Settings (Stats removed - now in dashboard)
        st.markdown("---")
        st.markdown("Settings")
        
        if st.button(" Profile", use_container_width=True, key="nav_profile"):
            st.session_state.page = 'profile'
            save_session_to_storage()
            st.rerun()
        
        
        
        # Logout
        st.markdown("---")
        if st.button(" Logout", use_container_width=True):
            st.query_params.clear()
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()


# ============================================
# MAIN APPLICATION
# ============================================
def main():
    """Main application entry point"""
    if not st.session_state.user:
        login_page()
        return
    
    # Render sidebar
    render_sidebar()
    
    # Route to appropriate page
    page = st.session_state.get('page', 'dashboard')
    
    if page == 'dashboard':
        display_dashboard()
    elif page == 'my_goals':
        display_my_goals()
    elif page == 'view_all_goals':
        display_view_all_goals()
    elif page == 'quarters':
        display_quarter_selection()
    elif page == 'months':
        display_month_selection()
    elif page == 'month_goals':
        display_month_goals()
    elif page == 'employees':
        display_employees_page()
    elif page == 'employee_goals':
        display_employee_goals()
    elif page == 'employee_quarters':
        display_quarter_selection()
    elif page == 'employee_months':
        display_month_selection()
    elif page == 'employee_month_goals':
        display_month_goals()
    elif page == 'employee_management':
        display_employee_management()
    elif page == 'hr_info':
        display_hr_info()
    elif page == 'analytics':  # ADD THIS
        display_analytics_page()
    elif page == 'feedback_history':
        display_feedback_history()
    elif page == 'profile':
        display_profile()
    elif page == 'permissions':
        display_permissions()
    else:
        # Default to dashboard
        st.session_state.page = 'dashboard'
        st.rerun()


if __name__ == "__main__":
    main()
